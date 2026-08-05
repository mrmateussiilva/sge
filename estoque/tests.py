from datetime import date
from decimal import Decimal

from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth.models import Group, User
import json
from io import BytesIO, StringIO

import openpyxl

from .models import Categoria, FechamentoMensal, Fornecedor, HistoricoPreco, ItemFechamento, LogAcao, Movimentacao, Produto
from .services.estoque_metrics import agrupar_quantidade_por_unidade
from .services.estoque_status import BAIXO, NORMAL, SEM_MINIMO, ZERADO, classificar_estoque, filtro_baixo, filtro_zerado
from .services.estoque_valuation import calcular_valor_estoque
from .services.units import decimal_br


class LoginTemplateTestCase(TestCase):
    def test_login_renderiza_com_next(self):
        response = self.client.get(f"{reverse('login')}?next=/produtos/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Acessar o SGE')
        self.assertContains(response, 'name="next" value="/produtos/"')
        self.assertContains(response, 'autocomplete="username"')
        self.assertContains(response, 'autocomplete="current-password"')

    def test_login_invalido_exibe_mensagem_generica_e_preserva_usuario(self):
        response = self.client.post(
            reverse('login'),
            data={
                'username': 'operador',
                'password': 'senha-incorreta',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Usuário ou senha inválidos.')
        self.assertContains(response, 'value="operador"')


class MovimentacaoTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='password123')
        self.client.login(username='testuser', password='password123')
        self.produto = Produto.objects.create(
            descricao='PRODUTO TESTE',
            tipo_produto='OUTRO',
            quantidade_base=Decimal('10.00'),
            preco_custo=Decimal('5.50'),
            preco_venda=Decimal('10.00'),
        )

    def test_registrar_entrada_com_quantidade_string(self):
        response = self.client.post(
            reverse('registrar_movimentacao'),
            data=json.dumps({
                'produto_id': self.produto.id,
                'tipo': 'ENTRADA',
                'quantidade': '4.00',
                'observacao': '',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])
        self.produto.refresh_from_db()
        self.assertEqual(self.produto.quantidade_base, Decimal('14.00'))

    def test_registrar_movimentacao_com_quantidade_invalida_retorna_400(self):
        response = self.client.post(
            reverse('registrar_movimentacao'),
            data=json.dumps({
                'produto_id': self.produto.id,
                'tipo': 'ENTRADA',
                'quantidade': '',
                'observacao': '',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()['ok'])
        self.produto.refresh_from_db()
        self.assertEqual(self.produto.quantidade_base, Decimal('10.00'))

    def test_registrar_saida_sem_saldo_retorna_400_com_erro(self):
        response = self.client.post(
            reverse('registrar_movimentacao'),
            data=json.dumps({
                'produto_id': self.produto.id,
                'tipo': 'SAIDA',
                'quantidade': '15.00',
                'observacao': '',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        data = response.json()
        self.assertFalse(data['ok'])
        self.assertIn('Saldo insuficiente', data['erro'])
        self.produto.refresh_from_db()
        self.assertEqual(self.produto.quantidade_base, Decimal('10.00'))

    def test_paginas_operacionais_renderizam(self):
        for name in (
            'dashboard',
            'lista_produtos',
            'registrar_movimentacao',
            'cadastrar_produto',
            'relatorio_mensal',
            'lista_fornecedores',
            'lista_categorias',
            'lista_fechamentos',
            'log_acoes',
            'lista_usuarios',
        ):
            with self.subTest(name=name):
                response = self.client.get(reverse(name))
                self.assertEqual(response.status_code, 200)

    def test_cadastrar_produto_com_campos_minimos(self):
        response = self.client.post(
            reverse('cadastrar_produto'),
            data=json.dumps({
                'tipo_produto': 'OUTRO',
                'unidade_medida': 'UN',
                'descricao': 'PRODUTO MINIMO',
                'quantidade_base': 0,
                'preco_custo': 0,
                'preco_venda': 0,
                'estoque_minimo': 0,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])
        self.assertTrue(Produto.objects.filter(descricao='PRODUTO MINIMO').exists())


class DominioEstoqueTestCase(TestCase):
    def criar_produto(self, **kwargs):
        defaults = {
            'descricao': 'PRODUTO',
            'tipo_produto': 'OUTRO',
            'unidade_medida': 'UN',
            'quantidade_base': Decimal('0.00'),
        }
        defaults.update(kwargs)
        return Produto.objects.create(**defaults)

    def test_unidade_base_e_formatacao_por_tipo(self):
        papel = self.criar_produto(descricao='PAPEL', tipo_produto='PAPEL', unidade_medida='UN', quantidade_base=Decimal('3000.00'))
        tecido = self.criar_produto(descricao='TECIDO', tipo_produto='TECIDO', unidade_medida='UN', quantidade_base=Decimal('1250.50'))
        tinta = self.criar_produto(descricao='TINTA', tipo_produto='TINTA', quantidade_base=Decimal('45.30'))
        unitario = self.criar_produto(descricao='UNITARIO', tipo_produto='OUTRO', unidade_medida='UN', quantidade_base=Decimal('18.00'))

        self.assertEqual(papel.unidade_simbolo, 'm')
        self.assertEqual(tecido.unidade_simbolo, 'm')
        self.assertEqual(tinta.unidade_simbolo, 'L')
        self.assertEqual(unitario.unidade_simbolo, 'un')
        self.assertEqual(papel.quantidade_formatada, '3.000,00 m')
        self.assertEqual(tinta.quantidade_formatada, '45,30 L')
        self.assertEqual(unitario.quantidade_formatada, '18,00 un')
        self.assertEqual(decimal_br(Decimal('1234.5')), '1.234,50')

    def test_estimativa_de_rolos_nao_altera_saldo_real(self):
        produto = self.criar_produto(
            tipo_produto='PAPEL',
            quantidade_base=Decimal('1000.00'),
            metros_por_rolo=Decimal('250.00'),
        )

        self.assertEqual(produto.quantidade_rolos_estimada, Decimal('4.00'))
        self.assertEqual(produto.quantidade_base, Decimal('1000.00'))

    def test_agregacao_de_movimentacoes_e_por_unidade(self):
        user = User.objects.create_user(username='movuser', password='password123')
        papel = self.criar_produto(descricao='PAPEL', tipo_produto='PAPEL')
        tinta = self.criar_produto(descricao='TINTA', tipo_produto='TINTA')
        outro = self.criar_produto(descricao='OUTRO', tipo_produto='OUTRO', unidade_medida='UN')
        movs = [
            Movimentacao.objects.create(produto=papel, usuario=user, tipo='ENTRADA', quantidade=Decimal('3000.00')),
            Movimentacao.objects.create(produto=tinta, usuario=user, tipo='ENTRADA', quantidade=Decimal('45.00')),
            Movimentacao.objects.create(produto=outro, usuario=user, tipo='ENTRADA', quantidade=Decimal('10.00')),
        ]

        totais = agrupar_quantidade_por_unidade(movs)
        self.assertEqual(totais['M'], Decimal('3000.00'))
        self.assertEqual(totais['L'], Decimal('45.00'))
        self.assertEqual(totais['UN'], Decimal('10.00'))

    def test_valuation_diferencia_zero_explicito_de_custo_ausente(self):
        vazio = calcular_valor_estoque([])
        self.assertTrue(vazio.calculo_completo)
        self.assertEqual(vazio.valor_conhecido, Decimal('0.00'))

        com_custo = self.criar_produto(quantidade_base=Decimal('10.00'), preco_custo=Decimal('5.00'))
        custo_zero = self.criar_produto(quantidade_base=Decimal('3.00'), preco_custo=Decimal('0.00'))
        sem_custo = self.criar_produto(quantidade_base=Decimal('7.00'), preco_custo=None)

        valuation = calcular_valor_estoque([com_custo, custo_zero, sem_custo])
        self.assertEqual(valuation.valor_conhecido, Decimal('50.00'))
        self.assertEqual(valuation.total_produtos_com_saldo, 3)
        self.assertEqual(valuation.produtos_com_custo, 2)
        self.assertEqual(valuation.produtos_sem_custo, 1)
        self.assertFalse(valuation.calculo_completo)

    def test_classificacao_de_estoque_e_querysets_disjuntos(self):
        casos = [
            (Decimal('0'), Decimal('0'), ZERADO),
            (Decimal('0'), Decimal('10'), ZERADO),
            (Decimal('5'), Decimal('10'), BAIXO),
            (Decimal('10'), Decimal('10'), BAIXO),
            (Decimal('11'), Decimal('10'), NORMAL),
            (Decimal('5'), Decimal('0'), SEM_MINIMO),
            (Decimal('5'), None, SEM_MINIMO),
        ]
        for idx, (saldo, minimo, esperado) in enumerate(casos):
            produto = self.criar_produto(descricao=f'P{idx}', quantidade_base=saldo, estoque_minimo=minimo)
            self.assertEqual(classificar_estoque(produto).codigo, esperado)

        baixos = set(Produto.objects.filter(filtro_baixo()).values_list('id', flat=True))
        zerados = set(Produto.objects.filter(filtro_zerado()).values_list('id', flat=True))
        self.assertTrue(baixos.isdisjoint(zerados))

    def test_lista_produtos_explica_contadores_filtrados(self):
        user = User.objects.create_user(username='listuser', password='password123')
        self.client.login(username='listuser', password='password123')
        categoria = Categoria.objects.create(nome='Tecidos')
        fornecedor = Fornecedor.objects.create(nome='Fornecedor A')
        self.criar_produto(descricao='TECIDO AZUL', tipo_produto='TECIDO', categoria=categoria, fornecedor=fornecedor, quantidade_base=Decimal('5'), estoque_minimo=Decimal('10'))
        self.criar_produto(descricao='TINTA CYAN', tipo_produto='TINTA', quantidade_base=Decimal('0'), estoque_minimo=Decimal('10'))

        response = self.client.get(reverse('lista_produtos'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Produtos nesta categoria')
        self.assertContains(response, 'Baixo nesta categoria')
        self.assertContains(response, 'Zerado nesta categoria')


class CategoriaHtmxTestCase(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='categoriasuser', password='password123')
        self.client.force_login(self.user)
        self.tecidos = Categoria.objects.create(
            nome='Tecidos',
            descricao='Materiais têxteis',
            cor='#112233',
        )
        self.tintas = Categoria.objects.create(
            nome='Tintas',
            descricao='Tintas de impressão',
            cor='#445566',
        )

    def test_pagina_renderiza_sem_vue_ou_json_embutido(self):
        response = self.client.get(reverse('lista_categorias'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="categorias-resultados"')
        self.assertContains(response, 'hx-get="/categorias/nova/"')
        self.assertNotContains(response, 'createApp')
        self.assertNotContains(response, 'categorias_json')

    def test_busca_htmx_retorna_somente_partial_filtrado(self):
        response = self.client.get(
            reverse('lista_categorias'),
            {'q': 'tecidos'},
            HTTP_HX_REQUEST='true',
            HTTP_HX_TARGET='categorias-resultados',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Tecidos')
        self.assertNotContains(response, 'Tintas')
        self.assertNotContains(response, '<h1>Categorias</h1>', html=True)

    def test_modal_htmx_de_edicao_vem_preenchido(self):
        response = self.client.get(
            reverse('editar_categoria', args=[self.tecidos.id]),
            {'q': 'tec'},
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-sge-modal-autoshow')
        self.assertContains(response, 'value="Tecidos"')
        self.assertContains(response, 'name="q" value="tec"')

    def test_criar_e_editar_categoria_com_htmx(self):
        criacao = self.client.post(
            reverse('criar_categoria'),
            data={'nome': 'Papéis', 'descricao': 'Papéis especiais', 'cor': '#abcdef'},
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(criacao.status_code, 200)
        self.assertContains(criacao, 'Papéis')
        self.assertIn('sge:modal-close', criacao['HX-Trigger-After-Swap'])
        categoria = Categoria.objects.get(nome='Papéis')
        self.assertTrue(LogAcao.objects.filter(
            acao='CRIAR', modelo='Categoria', objeto_id=categoria.id,
        ).exists())

        edicao = self.client.post(
            reverse('editar_categoria', args=[categoria.id]),
            data={'nome': 'Papéis premium', 'descricao': '', 'cor': '#fedcba'},
            HTTP_HX_REQUEST='true',
        )
        self.assertEqual(edicao.status_code, 200)
        categoria.refresh_from_db()
        self.assertEqual(categoria.nome, 'Papéis premium')
        self.assertEqual(categoria.cor, '#fedcba')

    def test_validacao_htmx_permanece_no_modal(self):
        response = self.client.post(
            reverse('criar_categoria'),
            data={'nome': '', 'cor': '#123456'},
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['HX-Retarget'], '#categoria-form-feedback')
        self.assertContains(response, 'Nome é obrigatório')

    def test_admin_exclui_categoria_com_htmx(self):
        admin = User.objects.create_superuser(username='admincategoria', password='password123')
        self.client.force_login(admin)

        response = self.client.post(
            reverse('excluir_categoria', args=[self.tintas.id]),
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="categorias-resultados"')
        self.assertNotContains(response, 'Tintas')
        self.assertIn('sge:feedback', response['HX-Trigger-After-Swap'])
        self.assertFalse(Categoria.objects.filter(pk=self.tintas.id).exists())


class FornecedorHtmxTestCase(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='fornecedoresuser', password='password123')
        self.client.force_login(self.user)
        self.alpha = Fornecedor.objects.create(
            nome='Fornecedor Alpha',
            cnpj='11.111.111/0001-11',
            email='alpha@example.com',
            telefone='(11) 11111-1111',
            observacao='Fornecedor principal',
        )
        self.beta = Fornecedor.objects.create(
            nome='Fornecedor Beta',
            cnpj='22.222.222/0001-22',
            email='beta@example.com',
        )

    def test_pagina_renderiza_sem_vue_ou_json_embutido(self):
        response = self.client.get(reverse('lista_fornecedores'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="fornecedores-resultados"')
        self.assertContains(response, 'hx-get="/fornecedores/novo/"')
        self.assertNotContains(response, 'createApp')
        self.assertNotContains(response, 'fornecedores_json')

    def test_busca_htmx_retorna_somente_partial_filtrado(self):
        response = self.client.get(
            reverse('lista_fornecedores'),
            {'q': 'alpha'},
            HTTP_HX_REQUEST='true',
            HTTP_HX_TARGET='fornecedores-resultados',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Fornecedor Alpha')
        self.assertNotContains(response, 'Fornecedor Beta')
        self.assertNotContains(response, '<h1>Fornecedores</h1>', html=True)

    def test_modal_htmx_de_edicao_vem_preenchido(self):
        response = self.client.get(
            reverse('editar_fornecedor', args=[self.alpha.id]),
            {'q': 'alpha'},
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-sge-modal-autoshow')
        self.assertContains(response, 'value="Fornecedor Alpha"')
        self.assertContains(response, 'Fornecedor principal')
        self.assertContains(response, 'name="q" value="alpha"')

    def test_criar_e_editar_fornecedor_com_htmx(self):
        criacao = self.client.post(
            reverse('criar_fornecedor'),
            data={
                'nome': 'Fornecedor Gama',
                'cnpj': '33.333.333/0001-33',
                'email': 'gama@example.com',
                'telefone': '(33) 33333-3333',
                'observacao': 'Criado via HTMX',
            },
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(criacao.status_code, 200)
        self.assertContains(criacao, 'Fornecedor Gama')
        self.assertIn('sge:modal-close', criacao['HX-Trigger-After-Swap'])
        fornecedor = Fornecedor.objects.get(nome='Fornecedor Gama')
        self.assertTrue(LogAcao.objects.filter(
            acao='CRIAR', modelo='Fornecedor', objeto_id=fornecedor.id,
        ).exists())

        edicao = self.client.post(
            reverse('editar_fornecedor', args=[fornecedor.id]),
            data={
                'nome': 'Fornecedor Gama Atualizado',
                'cnpj': fornecedor.cnpj,
                'email': 'novo@example.com',
                'telefone': '',
                'observacao': '',
            },
            HTTP_HX_REQUEST='true',
        )
        self.assertEqual(edicao.status_code, 200)
        fornecedor.refresh_from_db()
        self.assertEqual(fornecedor.nome, 'Fornecedor Gama Atualizado')
        self.assertEqual(fornecedor.email, 'novo@example.com')

    def test_validacao_htmx_permanece_no_modal(self):
        response = self.client.post(
            reverse('criar_fornecedor'),
            data={'nome': 'E-mail inválido', 'email': 'email-invalido'},
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['HX-Retarget'], '#fornecedor-form-feedback')
        self.assertContains(response, 'Informe um endereço de email válido')

    def test_admin_exclui_fornecedor_com_htmx(self):
        admin = User.objects.create_superuser(username='adminfornecedor', password='password123')
        self.client.force_login(admin)

        response = self.client.post(
            reverse('excluir_fornecedor', args=[self.beta.id]),
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="fornecedores-resultados"')
        self.assertNotContains(response, 'Fornecedor Beta')
        self.assertIn('sge:feedback', response['HX-Trigger-After-Swap'])
        self.assertFalse(Fornecedor.objects.filter(pk=self.beta.id).exists())


class UsernameDominioTestCase(TestCase):
    def test_criacao_normaliza_espacos_e_minusculas(self):
        user = User.objects.create_user(username='  Robson  ', password='password123')
        self.assertEqual(user.username, 'robson')

    def test_bloqueia_username_duplicado_case_insensitive(self):
        User.objects.create_user(username='robson', password='password123')
        with self.assertRaises(ValidationError):
            User.objects.create_user(username='Robson', password='password123')

    def test_edicao_do_proprio_usuario_sem_falso_conflito(self):
        user = User.objects.create_user(username='mateus', password='password123')
        user.is_staff = True
        user.save()
        user.refresh_from_db()
        self.assertEqual(user.username, 'mateus')

    def test_interface_de_usuarios_bloqueia_duplicado(self):
        admin = User.objects.create_superuser(username='admin', password='password123')
        User.objects.create_user(username='robson', password='password123')
        self.client.login(username='admin', password='password123')

        response = self.client.post(
            reverse('lista_usuarios'),
            data=json.dumps({'acao': 'criar', 'username': ' ROBSON ', 'password': 'password123'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()['ok'])

    def test_comando_audit_usernames_executa(self):
        out = StringIO()
        call_command('audit_usernames', stdout=out)
        self.assertIn('Nenhum username conflitante encontrado', out.getvalue())


class HistoricoPrecoTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='priceuser', password='password123')
        self.client.login(username='priceuser', password='password123')
        self.produto = Produto.objects.create(
            descricao='PRODUTO PRECO',
            tipo_produto='OUTRO',
            quantidade_base=Decimal('10.00'),
            preco_custo=Decimal('5.50'),
            preco_venda=Decimal('10.00'),
        )

    def test_salvamento_direto_cria_historico_quando_preco_muda(self):
        self.produto.preco_custo = Decimal('6.25')
        self.produto.save()

        historico = HistoricoPreco.objects.get(produto=self.produto)
        self.assertEqual(historico.preco_custo_antigo, Decimal('5.50'))
        self.assertEqual(historico.preco_custo_novo, Decimal('6.25'))
        self.assertIsNone(historico.preco_venda_antigo)
        self.assertIsNone(historico.preco_venda_novo)
        self.assertIsNone(historico.usuario)

    def test_salvamento_direto_sem_mudar_preco_nao_cria_historico(self):
        self.produto.descricao = 'PRODUTO PRECO EDITADO'
        self.produto.save()

        self.assertFalse(HistoricoPreco.objects.filter(produto=self.produto).exists())

    def test_editar_produto_cria_um_historico_com_usuario(self):
        response = self.client.post(
            reverse('editar_produto', args=[self.produto.id]),
            data=json.dumps({
                'tipo_produto': 'OUTRO',
                'unidade_medida': 'UN',
                'descricao': 'PRODUTO PRECO',
                'quantidade_base': '10.00',
                'preco_custo': '6.00',
                'preco_venda': '12.00',
                'estoque_minimo': '0.00',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])
        historicos = HistoricoPreco.objects.filter(produto=self.produto)
        self.assertEqual(historicos.count(), 1)
        historico = historicos.get()
        self.assertEqual(historico.preco_custo_antigo, Decimal('5.50'))
        self.assertEqual(historico.preco_custo_novo, Decimal('6.00'))
        self.assertEqual(historico.preco_venda_antigo, Decimal('10.00'))
        self.assertEqual(historico.preco_venda_novo, Decimal('12.00'))
        self.assertEqual(historico.usuario, self.user)


class FechamentoTestCase(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            password='testpassword'
        )
        self.client.login(username='testuser', password='testpassword')
        
        self.fornecedor = Fornecedor.objects.create(nome="FORNECEDOR TESTE")
        self.produto = Produto.objects.create(
            descricao="PRODUTO TESTE",
            tipo_produto="OUTRO",
            quantidade_base=10.0,
            preco_custo=5.50,
            preco_venda=10.00,
            fornecedor=self.fornecedor
        )

    def test_realizar_e_listar_fechamentos(self):
        response = self.client.post(
            reverse('realizar_fechamento'),
            data=json.dumps({
                'data_inicio': '2026-06-01',
                'data_fim': '2026-06-30',
                'observacao': 'Fechamento de teste'
            }),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['ok'])

        # Check database records
        self.assertEqual(FechamentoMensal.objects.count(), 1)
        fechamento = FechamentoMensal.objects.first()
        self.assertEqual(fechamento.data_inicio, date(2026, 6, 1))
        self.assertEqual(fechamento.data_fim, date(2026, 6, 30))
        self.assertEqual(fechamento.referencia_mes_ano, '06/2026')
        self.assertEqual(fechamento.itens.count(), 1)
        item = fechamento.itens.first()
        self.assertEqual(item.descricao, 'PRODUTO TESTE')
        self.assertEqual(item.quantidade, 10.0)
        self.assertEqual(item.tipo_produto, 'OUTRO')
        self.assertEqual(item.unidade_medida, 'UN')
        self.assertEqual(item.fornecedor_nome, 'FORNECEDOR TESTE')

        response = self.client.get(reverse('lista_fechamentos'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '01/06/2026 a 30/06/2026')
        self.assertContains(response, 'hx-post="/fechamentos/novo/"')
        self.assertNotContains(response, 'createApp')
        self.assertNotContains(response, 'fechamentos_json')

    def test_revisao_htmx_retorna_fragmento_html(self):
        response = self.client.get(
            reverse('revisar_fechamento'),
            {'data_inicio': '2026-10-01', 'data_fim': '2026-10-31'},
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Fechamento de 01/10/2026 a 31/10/2026')
        self.assertContains(response, 'Valor conhecido')
        self.assertNotContains(response, '"ok"')

    def test_criacao_htmx_aceita_formulario_e_redireciona(self):
        response = self.client.post(
            reverse('realizar_fechamento'),
            data={
                'data_inicio': '2026-11-01',
                'data_fim': '2026-11-30',
                'observacao': 'Criado com HTMX',
            },
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['HX-Redirect'], reverse('lista_fechamentos'))
        self.assertTrue(FechamentoMensal.objects.filter(
            data_inicio=date(2026, 11, 1),
            data_fim=date(2026, 11, 30),
            observacao='Criado com HTMX',
        ).exists())

    def test_exportar_fechamento_xlsx(self):
        # Create closure
        fechamento = FechamentoMensal.objects.create(
            data_inicio=date(2026, 6, 1),
            data_fim=date(2026, 6, 30),
            usuario=self.user
        )
        ItemFechamento.objects.create(
            fechamento=fechamento,
            produto=self.produto,
            descricao=self.produto.descricao,
            tipo_produto=self.produto.tipo_produto,
            unidade_medida='UN',
            fornecedor_nome=self.fornecedor.nome,
            quantidade=self.produto.quantidade_base,
            preco_custo=self.produto.preco_custo,
            preco_venda=self.produto.preco_venda
        )

        response = self.client.get(reverse('exportar_fechamento_xlsx', args=[fechamento.id]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('fechamento_01-06-2026_a_30-06-2026.xlsx', response['Content-Disposition'])
        ws = openpyxl.load_workbook(BytesIO(response.content), data_only=False).active
        self.assertEqual(
            [ws.cell(row=4, column=col).value for col in range(1, 11)],
            [
                'Descrição do Material', 'Tipo', 'Categoria', 'Fornecedor',
                'Unid.', 'Quantidade', 'Preço Custo', 'Preço Venda',
                'Total Custo', 'Total Venda',
            ],
        )
        self.assertEqual(ws['A5'].value, 'PRODUTO TESTE')
        self.assertEqual(ws['D5'].value, 'FORNECEDOR TESTE')

    def test_snapshot_permanece_igual_apos_produto_mudar_ou_ser_excluido(self):
        response = self.client.post(
            reverse('realizar_fechamento'),
            data=json.dumps({'data_inicio': '2026-07-10', 'data_fim': '2026-07-20'}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        fechamento = FechamentoMensal.objects.get(pk=response.json()['id'])
        item = fechamento.itens.get()

        self.produto.descricao = 'PRODUTO ALTERADO'
        self.produto.quantidade_base = Decimal('999.00')
        self.produto.preco_custo = Decimal('99.00')
        self.produto.save()

        admin = User.objects.create_superuser(username='adminfechamento', password='password123')
        self.client.force_login(admin)
        exclusao = self.client.post(reverse('excluir_produto', args=[self.produto.id]))
        self.assertEqual(exclusao.status_code, 200)

        item.refresh_from_db()
        self.assertIsNone(item.produto)
        self.assertEqual(item.descricao, 'PRODUTO TESTE')
        self.assertEqual(item.quantidade, Decimal('10.00'))
        self.assertEqual(item.preco_custo, Decimal('5.50'))

        detail = self.client.get(reverse('detalhe_fechamento', args=[fechamento.id]))
        self.assertContains(detail, 'PRODUTO TESTE')
        self.assertNotContains(detail, 'PRODUTO ALTERADO')

    def test_periodo_invalido_e_duplicado_sao_bloqueados(self):
        invalido = self.client.post(
            reverse('realizar_fechamento'),
            data=json.dumps({'data_inicio': '2026-08-31', 'data_fim': '2026-08-01'}),
            content_type='application/json',
        )
        self.assertEqual(invalido.status_code, 400)

        payload = {'data_inicio': '2026-08-01', 'data_fim': '2026-08-31'}
        primeiro = self.client.post(
            reverse('realizar_fechamento'), data=json.dumps(payload), content_type='application/json'
        )
        duplicado = self.client.post(
            reverse('realizar_fechamento'), data=json.dumps(payload), content_type='application/json'
        )
        self.assertEqual(primeiro.status_code, 200)
        self.assertEqual(duplicado.status_code, 400)
        self.assertEqual(duplicado.json()['codigo'], 'FECHAMENTO_DUPLICADO')

    def test_admin_exclui_fechamento_e_pode_refazer_periodo(self):
        payload = {'data_inicio': '2026-09-01', 'data_fim': '2026-09-30'}
        criado = self.client.post(
            reverse('realizar_fechamento'),
            data=json.dumps(payload),
            content_type='application/json',
        )
        fechamento_id = criado.json()['id']

        sem_permissao = self.client.post(reverse('excluir_fechamento', args=[fechamento_id]))
        self.assertEqual(sem_permissao.status_code, 403)
        self.assertTrue(FechamentoMensal.objects.filter(pk=fechamento_id).exists())

        admin = User.objects.create_superuser(username='adminfreeze', password='password123')
        self.client.force_login(admin)
        via_get = self.client.get(reverse('excluir_fechamento', args=[fechamento_id]))
        self.assertEqual(via_get.status_code, 405)

        exclusao = self.client.post(reverse('excluir_fechamento', args=[fechamento_id]))
        self.assertEqual(exclusao.status_code, 200)
        self.assertFalse(FechamentoMensal.objects.filter(pk=fechamento_id).exists())
        self.assertFalse(ItemFechamento.objects.filter(fechamento_id=fechamento_id).exists())
        self.assertTrue(LogAcao.objects.filter(
            acao='EXCLUIR',
            modelo='FechamentoMensal',
            objeto_id=fechamento_id,
        ).exists())

        recriado = self.client.post(
            reverse('realizar_fechamento'),
            data=json.dumps(payload),
            content_type='application/json',
        )
        self.assertEqual(recriado.status_code, 200)

    def test_exclusao_htmx_atualiza_partial_da_lista(self):
        admin = User.objects.create_superuser(username='adminhtmx', password='password123')
        self.client.force_login(admin)
        fechamento = FechamentoMensal.objects.create(
            data_inicio=date(2026, 12, 1),
            data_fim=date(2026, 12, 31),
            usuario=admin,
        )

        response = self.client.post(
            reverse('excluir_fechamento', args=[fechamento.id]),
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="fechamentos-lista"')
        self.assertContains(response, 'Nenhum fechamento realizado')
        self.assertIn('sge:feedback', response['HX-Trigger-After-Swap'])
        self.assertFalse(FechamentoMensal.objects.filter(pk=fechamento.id).exists())

    def test_exportar_atual_xlsx(self):
        response = self.client.get(reverse('exportar_atual_xlsx'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


class FluxosOperacionaisTestCase(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(username='adminop', password='password123')
        self.operador = User.objects.create_user(username='operadorop', password='password123')
        self.fornecedor = Fornecedor.objects.create(nome='Fornecedor Operacional')
        self.produto = Produto.objects.create(
            descricao='TECIDO OPERACIONAL',
            tipo_produto='TECIDO',
            quantidade_base=Decimal('45.30'),
            estoque_minimo=Decimal('20.00'),
            preco_custo=Decimal('10.00'),
            fornecedor=self.fornecedor,
        )

    def test_movimentacao_exibe_saldo_unidade_minimo_fornecedor_e_status(self):
        self.client.login(username='operadorop', password='password123')
        response = self.client.get(reverse('registrar_movimentacao'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '45,30 m')
        self.assertContains(response, '20,00 m')
        self.assertContains(response, 'Fornecedor Operacional')
        self.assertContains(response, 'Quantidade em')

    def test_saida_maior_que_saldo_e_bloqueada_no_backend(self):
        self.client.login(username='operadorop', password='password123')
        response = self.client.post(
            reverse('registrar_movimentacao'),
            data=json.dumps({'produto_id': self.produto.id, 'tipo': 'SAIDA', 'quantidade': '99'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['codigo'], 'SALDO_INSUFICIENTE')
        self.produto.refresh_from_db()
        self.assertEqual(self.produto.quantidade_base, Decimal('45.30'))

    def test_quantidade_zero_e_negativa_sao_bloqueadas(self):
        self.client.login(username='operadorop', password='password123')
        for quantidade in ('0', '-1'):
            with self.subTest(quantidade=quantidade):
                response = self.client.post(
                    reverse('registrar_movimentacao'),
                    data=json.dumps({'produto_id': self.produto.id, 'tipo': 'ENTRADA', 'quantidade': quantidade}),
                    content_type='application/json',
                )
                self.assertEqual(response.status_code, 400)

    def test_saida_que_deixa_estoque_baixo_registra_status(self):
        self.client.login(username='operadorop', password='password123')
        response = self.client.post(
            reverse('registrar_movimentacao'),
            data=json.dumps({'produto_id': self.produto.id, 'tipo': 'SAIDA', 'quantidade': '30.30'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status_estoque'], 'BAIXO')
        self.produto.refresh_from_db()
        self.assertEqual(self.produto.quantidade_base, Decimal('15.00'))

    def test_exclusao_via_get_e_bloqueada(self):
        self.client.login(username='adminop', password='password123')
        response = self.client.get(reverse('excluir_produto', args=[self.produto.id]))
        self.assertEqual(response.status_code, 405)

    def test_usuario_sem_permissao_nao_exclui_produto(self):
        self.client.login(username='operadorop', password='password123')
        response = self.client.post(reverse('excluir_produto', args=[self.produto.id]))
        self.assertEqual(response.status_code, 403)

    def test_fornecedor_com_produto_vinculado_nao_exclui(self):
        self.client.login(username='adminop', password='password123')
        response = self.client.post(reverse('excluir_fornecedor', args=[self.fornecedor.id]))
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['codigo'], 'VINCULO_IMPEDITIVO')

    def test_exclusao_de_movimentacao_recalcula_saldo_e_log(self):
        self.client.login(username='adminop', password='password123')
        mov = Movimentacao.objects.create(
            produto=self.produto,
            usuario=self.admin,
            tipo='ENTRADA',
            quantidade=Decimal('10.00'),
        )
        self.produto.refresh_from_db()
        self.assertEqual(self.produto.quantidade_base, Decimal('55.30'))

        response = self.client.post(reverse('excluir_movimentacao', args=[mov.id]))

        self.assertEqual(response.status_code, 200)
        self.produto.refresh_from_db()
        self.assertEqual(self.produto.quantidade_base, Decimal('45.30'))
        self.assertTrue(LogAcao.objects.filter(acao='EXCLUIR', modelo='Movimentacao', objeto_id=mov.id).exists())

    def test_revisao_fechamento_e_valor_parcial(self):
        Produto.objects.create(descricao='SEM CUSTO', tipo_produto='OUTRO', quantidade_base=Decimal('5'), preco_custo=None)
        self.client.login(username='operadorop', password='password123')

        response = self.client.get(reverse('revisar_fechamento'), {
            'data_inicio': '2026-07-01',
            'data_fim': '2026-07-31',
        })

        self.assertEqual(response.status_code, 200)
        resumo = response.json()['resumo']
        self.assertEqual(resumo['periodo_formatado'], '01/07/2026 a 31/07/2026')
        self.assertEqual(resumo['produtos_sem_custo'], 1)
        self.assertFalse(resumo['calculo_completo'])

    def test_fechamento_duplicado_e_bloqueado_e_snapshot_preservado(self):
        self.client.login(username='operadorop', password='password123')
        payload = {'data_inicio': '2026-07-01', 'data_fim': '2026-07-31', 'observacao': 'teste'}
        first = self.client.post(reverse('realizar_fechamento'), data=json.dumps(payload), content_type='application/json')
        second = self.client.post(reverse('realizar_fechamento'), data=json.dumps(payload), content_type='application/json')

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 400)
        self.assertEqual(FechamentoMensal.objects.filter(
            data_inicio=date(2026, 7, 1),
            data_fim=date(2026, 7, 31),
        ).count(), 1)
        self.assertTrue(LogAcao.objects.filter(acao='CRIAR', modelo='FechamentoMensal').exists())

    def test_alteracao_de_perfil_requer_admin(self):
        grupo, _ = Group.objects.get_or_create(name='Operador')
        self.client.login(username='operadorop', password='password123')
        response = self.client.post(
            reverse('lista_usuarios'),
            data=json.dumps({'acao': 'grupo', 'user_id': self.operador.id, 'grupo_id': grupo.id}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 403)

    def test_alteracao_de_perfil_registra_log(self):
        grupo, _ = Group.objects.get_or_create(name='Operador')
        self.client.login(username='adminop', password='password123')
        response = self.client.post(
            reverse('lista_usuarios'),
            data=json.dumps({'acao': 'grupo', 'user_id': self.operador.id, 'grupo_id': grupo.id}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        self.operador.refresh_from_db()
        self.assertTrue(self.operador.is_staff)
        self.assertTrue(LogAcao.objects.filter(acao='EDITAR', modelo='User', objeto_id=self.operador.id).exists())

    def test_nao_remove_propria_administracao(self):
        self.client.login(username='adminop', password='password123')
        response = self.client.post(
            reverse('lista_usuarios'),
            data=json.dumps({'acao': 'grupo', 'user_id': self.admin.id, 'grupo_id': None}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 400)
        self.admin.refresh_from_db()
        self.assertTrue(self.admin.is_superuser)


class ConfiguracaoOmieTestCase(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(username='adminomie', password='password123')
        self.user = User.objects.create_user(username='useromie', password='password123')

    def test_salvar_configuracao_omie_como_admin(self):
        self.client.login(username='adminomie', password='password123')
        response = self.client.post(
            reverse('salvar_configuracao_omie'),
            data=json.dumps({
                'app_key': 'KEY_TESTE_123',
                'app_secret': 'SECRET_TESTE_456',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])

        from .models import ConfiguracaoOmie
        config = ConfiguracaoOmie.objects.first()
        self.assertIsNotNone(config)
        self.assertEqual(config.app_key, 'KEY_TESTE_123')
        self.assertEqual(config.app_secret, 'SECRET_TESTE_456')
        self.assertEqual(config.usuario, self.admin)

    def test_salvar_configuracao_omie_requer_admin(self):
        self.client.login(username='useromie', password='password123')
        response = self.client.post(
            reverse('salvar_configuracao_omie'),
            data=json.dumps({
                'app_key': 'KEY_TESTE',
                'app_secret': 'SECRET_TESTE',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)

    def test_omie_client_usa_credenciais_do_banco(self):
        from .models import ConfiguracaoOmie
        from .services.omie_client import OmieClient

        ConfiguracaoOmie.objects.create(
            app_key='KEY_DO_BANCO',
            app_secret='SECRET_DO_BANCO',
        )

        client = OmieClient()
        self.assertEqual(client.app_key, 'KEY_DO_BANCO')
        self.assertEqual(client.app_secret, 'SECRET_DO_BANCO')

    def test_importar_nota_omie_criando_novo_produto(self):
        self.client.login(username='adminomie', password='password123')
        n_cod = 99999
        payload = {
            'itens': [
                {
                    'cod_item_int': '101',
                    'produto_id': 'novo',
                    'criar_novo': True,
                    'novo_descricao': 'PAO FRANCES INTEGRAL',
                    'novo_tipo_produto': 'OUTRO',
                    'novo_unidade_medida': 'UN',
                    'novo_estoque_minimo': '50.00',
                    'quantidade': '100.00',
                    'valor_unitario': '0.75',
                    'descricao': 'PAO FRANCES INTEGRAL',
                }
            ],
            'fornecedor_nome': 'PANIFICADORA TESTE',
            'numero_nfe': '999',
            'cod_int_nota_ent': 'INT999',
        }

        response = self.client.post(
            reverse('importar_nota_omie', args=[n_cod]),
            data=json.dumps(payload),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])

        # Verificar se o produto foi criado
        from .models import ImportacaoNFe, Movimentacao, Produto
        prod = Produto.objects.get(descricao='PAO FRANCES INTEGRAL')
        self.assertEqual(prod.unidade_medida, 'UN')
        self.assertEqual(prod.preco_custo, Decimal('0.75'))
        self.assertEqual(prod.quantidade_base, Decimal('100.00'))

        # Verificar movimentacao
        mov = Movimentacao.objects.get(produto=prod)
        self.assertEqual(mov.tipo, 'ENTRADA')
        self.assertEqual(mov.quantidade, Decimal('100.00'))

        # Verificar idempotencia
        self.assertTrue(ImportacaoNFe.objects.filter(n_cod_nota_ent=n_cod).exists())

    def test_listar_notas_entrada_com_filtros_e_ordenacao(self):
        from unittest.mock import MagicMock
        from .services.omie_client import OmieClient

        client = OmieClient(app_key='K', app_secret='S')
        client._chamar = MagicMock(return_value={'notas': [], 'nTotPaginas': 1, 'nTotRegistros': 0})

        client.listar_notas_entrada(
            pagina=1,
            registros_por_pagina=20,
            cnpj_fornecedor='12345678000199',
            data_inicio='01/01/2026',
            data_fim='28/07/2026',
            ordenar_decrescente=True,
        )

        client._chamar.assert_called_once_with(
            'produtos/notaentrada/',
            'ListarNotaEnt',
            {
                'nPagina': 1,
                'nRegistrosPorPagina': 20,
                'cCnpjForn': '12345678000199',
                'dEmiInicial': '01/01/2026',
                'dEmiFinal': '28/07/2026',
            }
        )


class HTMXViewsTestCase(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(username='admin', password='password123')
        self.client = Client()
        self.client.login(username='admin', password='password123')
        self.fornecedor = Fornecedor.objects.create(nome='Fornecedor HTMX')
        self.produto = Produto.objects.create(
            descricao='PRODUTO HTMX TESTE',
            tipo_produto='PAPEL',
            unidade_medida='UN',
            quantidade_base=Decimal('50.00'),
            estoque_minimo=Decimal('10.00'),
            preco_custo=Decimal('15.00'),
            fornecedor=self.fornecedor,
        )

    def test_lista_produtos_renderiza_vue_com_dados(self):
        response = self.client.get(reverse('lista_produtos'), HTTP_HX_REQUEST='true')
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'estoque/lista.html')
        self.assertContains(response, 'id="app"')
        self.assertContains(response, 'createApp')
        self.assertContains(response, 'PRODUTO HTMX TESTE')
        self.assertContains(response, "tipo_produto\": \"PAPEL")

    def test_atualiza_estoque_htmx_retorna_cell_e_header(self):
        response = self.client.post(
            reverse('atualiza_estoque'),
            data={'id': self.produto.id, 'variacao': 1},
            HTTP_HX_REQUEST='true',
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'estoque/produtos/_quantidade_cell.html')
        self.assertEqual(response.headers.get('HX-Trigger'), 'estoqueAtualizado')

    def test_inline_edit_estoque_htmx(self):
        response = self.client.get(reverse('inline_edit_estoque', args=[self.produto.id]), HTTP_HX_REQUEST='true')
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'estoque/produtos/_inline_edit_form.html')

        post_response = self.client.post(
            reverse('inline_edit_estoque', args=[self.produto.id]),
            data={'quantidade_base': '60.00'},
            HTTP_HX_REQUEST='true',
        )
        self.assertEqual(post_response.status_code, 200)
        self.assertTemplateUsed(post_response, 'estoque/produtos/_quantidade_cell.html')
        self.assertEqual(post_response.headers.get('HX-Trigger'), 'estoqueAtualizado')

    def test_lista_ordens_htmx_retorna_partial(self):
        response = self.client.get(reverse('lista_ordens'), HTTP_HX_REQUEST='true')
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'estoque/ordens/_lista_resultados.html')

    def test_log_acoes_htmx_retorna_partial(self):
        response = self.client.get(reverse('log_acoes'), HTTP_HX_REQUEST='true')
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'estoque/logs/_lista_resultados.html')

    def test_info_produto_movimentacao_htmx(self):
        response = self.client.get(f"{reverse('info_produto_movimentacao')}?produto_id={self.produto.id}", HTTP_HX_REQUEST='true')
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'estoque/movimentacao/_produto_info.html')
        self.assertContains(response, '50,00 m')

    def test_registrar_movimentacao_htmx_retorna_tabela_e_header(self):
        response = self.client.post(
            reverse('registrar_movimentacao'),
            data={'produto_id': self.produto.id, 'tipo': 'ENTRADA', 'quantidade': '10.00', 'observacao': 'Teste HTMX'},
            HTTP_HX_REQUEST='true',
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'estoque/movimentacao/_historico_tabela.html')
        self.assertEqual(response.headers.get('HX-Trigger'), 'estoqueAtualizado')
