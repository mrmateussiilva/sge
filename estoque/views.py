import calendar
import csv
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Count, F, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.utils import timezone

from .log_utils import log_acao
from .models import Categoria, Fornecedor, HistoricoPreco, ItemOrdemCompra, LogAcao, Movimentacao, OrdemCompra, Produto, FechamentoMensal
from .services.estoque_metrics import agrupar_quantidade_por_unidade, serializar_totais_unidade, valor_por_tipo
from .services.estoque_status import filtro_baixo, filtro_normal, filtro_sem_minimo, filtro_zerado
from .services.estoque_valuation import calcular_valor_estoque
from .services.fechamentos import criar_fechamento_periodo, validar_periodo
from .services.units import UNIDADES, decimal_br, dinheiro_br, formatar_capacidade_embalagem, formatar_quantidade, unidade_base_codigo, unidade_info
from .services.usernames import validate_username_available

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def decimal_ou_none(value):
    if value in (None, ''):
        return None
    return Decimal(str(value))


def data_iso(value, nome_campo):
    try:
        return date.fromisoformat(str(value))
    except (TypeError, ValueError):
        raise ValidationError(f'{nome_campo} inválida. Use o formato AAAA-MM-DD.')


PERFIS_NEGOCIO = {
    'Admin': 'Administrador',
    'Gestor': 'Gestor de estoque',
    'Operador': 'Operador',
    'Leitura': 'Somente leitura',
    'Visualizador': 'Somente leitura',
}


def json_ok(**kwargs):
    return JsonResponse({'ok': True, **kwargs})


def json_erro(mensagem, status=400, codigo='VALIDACAO'):
    return JsonResponse({'ok': False, 'erro': mensagem, 'codigo': codigo}, status=status)


def requisicao_htmx(request):
    return request.headers.get('HX-Request') == 'true'


def usuario_pode_alterar(request):
    return request.user.is_superuser


def exigir_admin_json(request):
    if not usuario_pode_alterar(request):
        return json_erro('Permissão negada.', status=403, codigo='PERMISSAO_NEGADA')
    return None


def produto_operacional_json(produto):
    minimo = produto.estoque_minimo
    return {
        'id': produto.id,
        'descricao': produto.descricao,
        'fornecedor': produto.fornecedor.nome if produto.fornecedor else '',
        'quantidade': float(produto.quantidade_base),
        'quantidade_formatada': produto.quantidade_formatada,
        'estoque_minimo': float(minimo) if minimo is not None else None,
        'estoque_minimo_formatado': formatar_quantidade(minimo, produto.unidade_base_codigo) if minimo is not None else 'Sem mínimo configurado',
        'unidade_codigo': produto.unidade_base_codigo,
        'unidade_simbolo': produto.unidade_simbolo,
        'unidade_nome': unidade_info(produto).plural,
        'status_estoque': produto.status_estoque,
    }


def resumo_fechamento(data_inicio, data_fim):
    validar_periodo(data_inicio, data_fim)
    produtos = list(Produto.objects.select_related('fornecedor').all())
    valuation = calcular_valor_estoque(produtos)
    sem_fornecedor = sum(1 for p in produtos if not p.fornecedor_id)
    return {
        'data_inicio': data_inicio.isoformat(),
        'data_fim': data_fim.isoformat(),
        'periodo_formatado': f'{data_inicio:%d/%m/%Y} a {data_fim:%d/%m/%Y}',
        'total_produtos': len(produtos),
        'produtos_zerados': Produto.objects.filter(filtro_zerado()).count(),
        'produtos_baixos': Produto.objects.filter(filtro_baixo()).count(),
        'produtos_sem_custo': valuation.produtos_sem_custo,
        'produtos_sem_fornecedor': sem_fornecedor,
        'valor_conhecido': float(valuation.valor_conhecido),
        'valor_conhecido_formatado': dinheiro_br(valuation.valor_conhecido),
        'calculo_completo': valuation.calculo_completo,
        'duplicado': FechamentoMensal.objects.filter(
            data_inicio=data_inicio,
            data_fim=data_fim,
        ).exists(),
    }


@login_required
def dashboard(request):
    produtos_base = Produto.objects.select_related('fornecedor').all()
    produtos_lista = list(produtos_base)
    total_itens = len(produtos_lista)
    estoque_zerado_count = Produto.objects.filter(filtro_zerado()).count()
    estoque_baixo = Produto.objects.filter(filtro_baixo()).select_related('fornecedor')
    valuation = calcular_valor_estoque(produtos_lista)
    ultimas_movimentacoes = Movimentacao.objects.select_related('produto', 'usuario').order_by('-data')[:5]

    hoje = timezone.now()
    ano_atual = hoje.year
    try:
        ano_selecionado = int(request.GET.get('ano', ano_atual))
    except ValueError:
        ano_selecionado = ano_atual

    # Obter anos que possuem movimentações para o seletor
    anos_disponiveis = list(Movimentacao.objects.dates('data', 'year', order='DESC'))
    anos = sorted(list(set([a.year for a in anos_disponiveis] + [ano_atual])), reverse=True)
    
    # Gerar dados mensais corretos para o ano selecionado
    meses_nomes = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    entradas_meses = []
    saidas_meses = []
    for m in range(1, 13):
        entradas = Movimentacao.objects.filter(
            tipo='ENTRADA', data__year=ano_selecionado, data__month=m
        ).count()
        saidas = Movimentacao.objects.filter(
            tipo='SAIDA', data__year=ano_selecionado, data__month=m
        ).count()
        entradas_meses.append(entradas)
        saidas_meses.append(saidas)

    if request.GET.get('ajax') == '1':
        return JsonResponse({
            'meses': meses_nomes,
            'entradas': entradas_meses,
            'saidas': saidas_meses,
        })

    valores_por_tipo = valor_por_tipo(produtos_lista)
    tipo_choices = dict(Produto.TIPO_PRODUTO_CHOICES)
    tipo_labels = [tipo_choices.get(tipo, tipo) for tipo in valores_por_tipo.keys()]
    tipo_data = [float(total) for total in valores_por_tipo.values()]

    valor_total = valuation.valor_conhecido

    return render(request, 'estoque/dashboard.html', {
        'total_itens': total_itens,
        'estoque_zerado_count': estoque_zerado_count,
        'valor_total': valor_total,
        'valuation': valuation,
        'valor_total_formatado': dinheiro_br(valor_total),
        'estoque_baixo': estoque_baixo,
        'ultimas_movimentacoes': ultimas_movimentacoes,
        'ultimos_logs': LogAcao.objects.select_related('usuario').all()[:5],
        'chart_meses': json.dumps(meses_nomes),
        'chart_entradas': json.dumps(entradas_meses),
        'chart_saidas': json.dumps(saidas_meses),
        'chart_tipo_labels': json.dumps(tipo_labels),
        'chart_tipo_data': json.dumps(tipo_data),
        'anos_disponiveis': anos,
        'ano_selecionado': ano_selecionado,
    })


@login_required
def lista_produtos(request):
    busca = (request.GET.get('busca') or request.GET.get('q') or '').strip()
    filtro_url = (request.GET.get('filtro') or '').lower()
    filtro_map = {'baixo': 'BAIXO', 'zerado': 'ZERADO', 'ok': 'OK'}
    filtro_estoque = request.GET.get('estoque') or filtro_map.get(filtro_url, 'TODOS')
    if filtro_estoque not in ('TODOS', 'BAIXO', 'ZERADO', 'OK'):
        filtro_estoque = 'TODOS'

    filtro_fornecedor = request.GET.get('fornecedor') or 'TODOS'
    aba_ativa = (request.GET.get('aba') or 'PAPEL').upper()
    abas_validas = ['PAPEL', 'TECIDO', 'TINTA', 'AVIAMENTO', 'OUTRO']
    if aba_ativa not in abas_validas:
        aba_ativa = 'PAPEL'

    ordem_coluna = request.GET.get('ordem') or 'descricao'
    ordem_direcao = request.GET.get('direcao') or 'asc'

    qs = Produto.objects.select_related('fornecedor', 'categoria').all()

    if busca:
        qs = qs.filter(
            Q(descricao__icontains=busca) | Q(fornecedor__nome__icontains=busca)
        )

    if filtro_fornecedor == 'SEM_FORNECEDOR':
        qs = qs.filter(fornecedor__isnull=True)
    elif filtro_fornecedor != 'TODOS':
        qs = qs.filter(fornecedor__nome=filtro_fornecedor)

    if filtro_estoque == 'ZERADO':
        qs = qs.filter(filtro_zerado())
    elif filtro_estoque == 'BAIXO':
        qs = qs.filter(filtro_baixo())
    elif filtro_estoque == 'OK':
        qs = qs.filter(filtro_normal() | filtro_sem_minimo())

    fornecedores_unicos = list(
        Fornecedor.objects.filter(produto__isnull=False)
        .values_list('nome', flat=True).distinct().order_by('nome')
    )

    produtos_filtrados = list(qs)
    total_filtrado = len(produtos_filtrados)
    total_produtos = Produto.objects.count()

    tabs = [
        {'key': 'PAPEL',     'label': 'Papel',      'icon': 'bi bi-file-earmark-text'},
        {'key': 'TECIDO',    'label': 'Tecido',     'icon': 'bi bi-grid-3x3-gap'},
        {'key': 'TINTA',     'label': 'Tinta',      'icon': 'bi bi-droplet-half'},
        {'key': 'AVIAMENTO', 'label': 'Aviamentos', 'icon': 'bi bi-tools'},
        {'key': 'OUTRO',     'label': 'Outros',     'icon': 'bi bi-three-dots'},
    ]

    produtos_por_aba = {t['key']: [] for t in tabs}
    for p in produtos_filtrados:
        if p.tipo_produto in produtos_por_aba:
            produtos_por_aba[p.tipo_produto].append(p)
        else:
            produtos_por_aba['OUTRO'].append(p)

    tem_filtros_ativos = bool(busca or filtro_estoque != 'TODOS' or filtro_fornecedor != 'TODOS')
    if tem_filtros_ativos and len(produtos_por_aba[aba_ativa]) == 0:
        for t in tabs:
            if len(produtos_por_aba[t['key']]) > 0:
                aba_ativa = t['key']
                break

    all_produtos = list(Produto.objects.all())
    critico_por_tipo = {}
    for p in all_produtos:
        if p.status_estoque in ('ZERADO', 'BAIXO'):
            critico_por_tipo[p.tipo_produto] = True

    tabs_data = []
    for t in tabs:
        k = t['key']
        tabs_data.append({
            'key': k,
            'label': t['label'],
            'icon': t['icon'],
            'count': len(produtos_por_aba[k]),
            'has_critical': critico_por_tipo.get(k, False),
        })

    produtos_aba = produtos_por_aba[aba_ativa]

    def get_sort_key(p):
        if ordem_coluna == 'fornecedor':
            val = p.fornecedor.nome if p.fornecedor else ''
        elif ordem_coluna == 'metros_por_rolo':
            val = float(p.metros_por_rolo or 0)
        elif ordem_coluna == 'quantidade':
            val = float(p.quantidade_base)
        elif ordem_coluna == 'preco_custo':
            val = float(p.preco_custo or 0)
        else:
            val = p.descricao or ''
        return val

    reverse_sort = (ordem_direcao == 'desc')
    if ordem_coluna in ('quantidade', 'metros_por_rolo', 'preco_custo'):
        produtos_aba.sort(key=get_sort_key, reverse=reverse_sort)
    else:
        produtos_aba.sort(key=lambda p: str(get_sort_key(p)).lower(), reverse=reverse_sort)

    valor_custo = Decimal('0.00')
    sem_custo = 0
    baixos = 0
    zerados = 0

    for p in produtos_aba:
        st = p.status_estoque
        if st == 'ZERADO':
            zerados += 1
        elif st == 'BAIXO':
            baixos += 1

        if p.preco_custo is not None:
            valor_custo += (p.quantidade_base * p.preco_custo)
        elif p.quantidade_base > 0:
            sem_custo += 1

    paginator = Paginator(produtos_aba, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    extra_params = f"&aba={aba_ativa}&busca={busca}&estoque={filtro_estoque}&fornecedor={filtro_fornecedor}&ordem={ordem_coluna}&direcao={ordem_direcao}"

    resumo_aba = {
        'total_itens': len(produtos_aba),
        'valor_custo': valor_custo,
        'sem_custo': sem_custo,
        'baixos': baixos,
        'zerados': zerados,
    }

    context = {
        'page_obj': page_obj,
        'produtos': page_obj,
        'resumo_aba': resumo_aba,
        'tabs': tabs_data,
        'aba_ativa': aba_ativa,
        'busca': busca,
        'filtro_estoque': filtro_estoque,
        'filtro_fornecedor': filtro_fornecedor,
        'fornecedores_unicos': fornecedores_unicos,
        'ordem': ordem_coluna,
        'direcao': ordem_direcao,
        'total_filtrado': total_filtrado,
        'total_produtos': total_produtos,
        'tem_filtros_ativos': tem_filtros_ativos,
        'extra_params': extra_params,
    }

    if request.headers.get('HX-Request'):
        return render(request, 'estoque/produtos/_lista_resultados.html', context)
    return render(request, 'estoque/lista.html', context)


@login_required
def atualiza_estoque(request):
    if request.method == 'POST':
        try:
            if request.content_type == 'application/json':
                data = json.loads(request.body)
            else:
                data = request.POST

            produto_id = data.get('id')
            variacao = Decimal(str(data.get('variacao', 0)))
            produto = Produto.objects.get(id=produto_id)

            if variacao != 0:
                tipo = 'ENTRADA' if variacao > 0 else 'SAIDA'
                quantidade = abs(variacao)
                with transaction.atomic():
                    Movimentacao.objects.create(
                        produto=produto,
                        usuario=request.user,
                        tipo=tipo,
                        quantidade=quantidade,
                        observacao='Ajuste rápido de estoque',
                    )
                produto.refresh_from_db()

            if request.headers.get('HX-Request'):
                response = render(request, 'estoque/produtos/_quantidade_cell.html', {'p': produto})
                response['HX-Trigger'] = 'estoqueAtualizado'
                return response

            return json_ok(
                nova_quantidade=float(produto.quantidade_base),
                nova_quantidade_formatada=produto.quantidade_formatada,
                status_estoque=produto.status_estoque,
            )
        except (InvalidOperation, TypeError, ValueError):
            if request.headers.get('HX-Request'):
                return HttpResponse('Quantidade inválida.', status=400)
            return json_erro('Quantidade inválida.')
        except Produto.DoesNotExist:
            if request.headers.get('HX-Request'):
                return HttpResponse('Produto não encontrado.', status=404)
            return json_erro('Produto não encontrado.', status=404)
        except ValidationError as e:
            msg = '; '.join(e.messages)
            if request.headers.get('HX-Request'):
                return HttpResponse(msg, status=400)
            return json_erro(msg, codigo='SALDO_INSUFICIENTE')
    return json_erro('Método não permitido.', status=405)


@login_required
def inline_edit_estoque(request, id):
    produto = get_object_or_404(Produto, pk=id)

    if request.GET.get('cancel'):
        return render(request, 'estoque/produtos/_quantidade_cell.html', {'p': produto})

    if request.method == 'POST':
        try:
            nova_qtd = Decimal(str(request.POST.get('quantidade_base', produto.quantidade_base)))
            variacao = nova_qtd - produto.quantidade_base
            if variacao != 0:
                tipo = 'ENTRADA' if variacao > 0 else 'SAIDA'
                with transaction.atomic():
                    Movimentacao.objects.create(
                        produto=produto,
                        usuario=request.user,
                        tipo=tipo,
                        quantidade=abs(variacao),
                        observacao='Edição inline de quantidade',
                    )
                produto.refresh_from_db()
            response = render(request, 'estoque/produtos/_quantidade_cell.html', {'p': produto})
            response['HX-Trigger'] = 'estoqueAtualizado'
            return response
        except (InvalidOperation, TypeError, ValueError):
            return HttpResponse('Quantidade inválida.', status=400)
        except ValidationError as e:
            return HttpResponse('; '.join(e.messages), status=400)

    return render(request, 'estoque/produtos/_inline_edit_form.html', {'p': produto})


@login_required
def registrar_movimentacao(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            produto = Produto.objects.get(id=data['produto_id'])
            tipo = data['tipo']
            quantidade = Decimal(str(data['quantidade']))
            if quantidade <= 0:
                return json_erro('Quantidade deve ser maior que zero.')
            if tipo == 'SAIDA' and produto.quantidade_base < quantidade:
                return json_erro('Saldo insuficiente para esta saída.', codigo='SALDO_INSUFICIENTE')
            Movimentacao.objects.create(
                produto=produto,
                usuario=request.user,
                tipo=tipo,
                quantidade=quantidade,
                observacao=data.get('observacao', ''),
            )
            produto.refresh_from_db()
            log_acao(request.user, tipo, f'{tipo} de {quantidade} {produto.unidade_simbolo} de {produto.descricao}', 'Movimentacao')
            return json_ok(
                mensagem='Movimentação registrada com sucesso.',
                saldo_atual=produto.quantidade_formatada,
                status_estoque=produto.status_estoque,
            )
        except json.JSONDecodeError:
            return json_erro('JSON inválido.')
        except KeyError as e:
            return json_erro(f'Campo obrigatório ausente: {e.args[0]}.')
        except (InvalidOperation, TypeError, ValueError):
            return json_erro('Quantidade inválida.')
        except Produto.DoesNotExist:
            return json_erro('Produto não encontrado.', status=404)
        except ValidationError as e:
            return json_erro('; '.join(e.messages), codigo='SALDO_INSUFICIENTE')
    produtos = [produto_operacional_json(p) for p in Produto.objects.select_related('fornecedor').all().order_by('descricao')]
    movimentacoes_qs = Movimentacao.objects.select_related('produto', 'usuario').order_by('-data')
    paginator = Paginator(movimentacoes_qs, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'estoque/movimentacao.html', {
        'produtos': produtos,
        'page_obj': page_obj,
        'movimentacoes': page_obj,
    })


@login_required
def exportar_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="relatorio_estoque.csv"'

    writer = csv.writer(response)
    writer.writerow(['Descricao', 'Tipo', 'Fornecedor', 'Unidade Base', 'Quantidade', 'Preco Custo', 'Preco Venda', 'Estoque Minimo'])
    produtos = Produto.objects.select_related('fornecedor').all()
    for p in produtos:
        writer.writerow([
            p.descricao,
            p.get_tipo_produto_display(),
            p.fornecedor.nome if p.fornecedor else '',
            p.unidade_simbolo,
            p.quantidade_formatada,
            p.preco_custo if p.preco_custo is not None else '',
            p.preco_venda if p.preco_venda is not None else '',
            p.estoque_minimo if p.estoque_minimo is not None else '',
        ])
    return response


@login_required
def cadastrar_produto(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        produto = Produto.objects.create(
            tipo_produto=data['tipo_produto'],
            descricao=data['descricao'],
            fornecedor_id=data.get('fornecedor_id') or None,
            quantidade_base=data.get('quantidade_base', 0),
            preco_custo=decimal_ou_none(data.get('preco_custo')),
            preco_venda=decimal_ou_none(data.get('preco_venda')),
            estoque_minimo=decimal_ou_none(data.get('estoque_minimo')),
            metros_por_rolo=data.get('metros_por_rolo') or None,
            tipo_tinta=data.get('tipo_tinta', 'N/A'),
            cor_tinta=data.get('cor_tinta', 'INCOLOR'),
            litros_por_vidro=data.get('litros_por_vidro') or None,
            unidade_medida=data.get('unidade_medida', 'UN'),
            categoria_id=data.get('categoria_id') or None,
        )
        log_acao(request.user, 'CRIAR', f'Cadastrou produto {produto.descricao}', 'Produto', produto.id)
        return JsonResponse({'ok': True, 'id': produto.id})
    fornecedores = Fornecedor.objects.all().values('id', 'nome')
    categorias = Categoria.objects.all().values('id', 'nome')
    return render(request, 'estoque/cadastrar_produto.html', {
        'fornecedores': list(fornecedores),
        'categorias': list(categorias),
    })


@login_required
def editar_produto(request, id):
    produto = get_object_or_404(Produto, id=id)
    if request.method == 'POST':
        data = json.loads(request.body)
        old_preco_custo = produto.preco_custo
        old_preco_venda = produto.preco_venda
        produto.tipo_produto = data['tipo_produto']
        produto.descricao = data['descricao']
        produto.fornecedor_id = data.get('fornecedor_id') or None
        produto.quantidade_base = data.get('quantidade_base', 0)
        produto.preco_custo = decimal_ou_none(data.get('preco_custo'))
        produto.preco_venda = decimal_ou_none(data.get('preco_venda'))
        produto.estoque_minimo = decimal_ou_none(data.get('estoque_minimo'))
        produto.metros_por_rolo = data.get('metros_por_rolo') or None
        produto.tipo_tinta = data.get('tipo_tinta', 'N/A')
        produto.cor_tinta = data.get('cor_tinta', 'INCOLOR')
        produto.litros_por_vidro = data.get('litros_por_vidro') or None
        produto.unidade_medida = data.get('unidade_medida', 'UN')
        produto.categoria_id = data.get('categoria_id') or None
        preco_custo_mudou = old_preco_custo != produto.preco_custo
        preco_venda_mudou = old_preco_venda != produto.preco_venda
        produto._historico_ja_salvo = True
        produto.save()
        if preco_custo_mudou or preco_venda_mudou:
            HistoricoPreco.objects.create(
                produto=produto,
                preco_custo_antigo=old_preco_custo if preco_custo_mudou else None,
                preco_custo_novo=produto.preco_custo if preco_custo_mudou else None,
                preco_venda_antigo=old_preco_venda if preco_venda_mudou else None,
                preco_venda_novo=produto.preco_venda if preco_venda_mudou else None,
                usuario=request.user,
            )
        log_acao(request.user, 'EDITAR', f'Editou produto {produto.descricao}', 'Produto', produto.id)
        return JsonResponse({'ok': True})
    fornecedores = Fornecedor.objects.all().values('id', 'nome')
    categorias = Categoria.objects.all().values('id', 'nome')
    return render(request, 'estoque/editar_produto.html', {
        'produto': produto,
        'fornecedores': list(fornecedores),
        'categorias': list(categorias),
        'produto_json': json.dumps({
            'id': produto.id,
            'tipo_produto': produto.tipo_produto,
            'descricao': produto.descricao,
            'fornecedor_id': produto.fornecedor_id or '',
            'categoria_id': produto.categoria_id or '',
            'quantidade_base': float(produto.quantidade_base) if produto.quantidade_base else '',
            'preco_custo': float(produto.preco_custo) if produto.preco_custo is not None else '',
            'preco_venda': float(produto.preco_venda) if produto.preco_venda is not None else '',
            'estoque_minimo': float(produto.estoque_minimo) if produto.estoque_minimo is not None else '',
            'metros_por_rolo': float(produto.metros_por_rolo) if produto.metros_por_rolo else None,
            'tipo_tinta': produto.tipo_tinta,
            'cor_tinta': produto.cor_tinta,
            'litros_por_vidro': float(produto.litros_por_vidro) if produto.litros_por_vidro else None,
            'unidade_medida': produto.unidade_medida or 'UN',
        }),
    })


@login_required
def excluir_produto(request, id):
    produto = get_object_or_404(Produto, id=id)
    if request.method == 'POST':
        perm_error = exigir_admin_json(request)
        if perm_error:
            return perm_error
        if produto.movimentacoes.exists():
            return json_erro('O produto não pode ser excluído porque possui movimentações vinculadas.', codigo='VINCULO_IMPEDITIVO')
        if ItemOrdemCompra.objects.filter(produto=produto).exists():
            return json_erro('O produto não pode ser excluído porque possui ordens de compra vinculadas.', codigo='VINCULO_IMPEDITIVO')
        descricao = produto.descricao
        produto.delete()
        log_acao(request.user, 'EXCLUIR', f'Excluiu produto {descricao}', 'Produto', id)
        return json_ok(mensagem='Produto excluído com sucesso.')
    return json_erro('Exclusão deve usar POST.', status=405)


@login_required
def excluir_movimentacao(request, id):
    mov = get_object_or_404(Movimentacao, id=id)
    if request.method == 'POST':
        perm_error = exigir_admin_json(request)
        if perm_error:
            return perm_error
        with transaction.atomic():
            produto = Produto.objects.select_for_update().get(pk=mov.produto.pk)
            if mov.tipo == 'ENTRADA':
                produto.quantidade_base -= mov.quantidade
            else:
                produto.quantidade_base += mov.quantidade
            produto.save()
            descricao = f'{mov.get_tipo_display()} de {mov.quantidade} de {mov.produto.descricao}'
            mov.delete()
        log_acao(request.user, 'EXCLUIR', f'Excluiu movimentacao: {descricao}', 'Movimentacao', id)
        return json_ok(mensagem='Movimentação excluída e saldo recalculado com sucesso.', nova_quantidade=produto.quantidade_formatada)
    return json_erro('Exclusão deve usar POST.', status=405)


@login_required
def detalhe_produto(request, id):
    produto = get_object_or_404(Produto, id=id)
    movimentacoes = Movimentacao.objects.filter(produto=produto).select_related('produto').order_by('-data')
    historico_precos = HistoricoPreco.objects.filter(produto=produto)[:20]
    lucro = (produto.preco_venda - produto.preco_custo) if produto.preco_venda is not None and produto.preco_custo is not None else None
    margem = (lucro / produto.preco_custo * 100) if lucro is not None and produto.preco_custo and produto.preco_custo > 0 else None
    return render(request, 'estoque/detalhe.html', {
        'produto': produto,
        'movimentacoes': movimentacoes,
        'historico_precos': historico_precos,
        'lucro': round(lucro, 2) if lucro is not None else None,
        'margem': round(margem, 1) if margem is not None else None,
    })


@login_required
def lista_ordens(request):
    busca = (request.GET.get('busca') or request.GET.get('q') or '').strip()
    status_selecionado = request.GET.get('status', '').strip()
    fornecedor_selecionado = request.GET.get('fornecedor', '').strip()

    qs = OrdemCompra.objects.select_related('fornecedor').all().order_by('-data_criacao')

    if busca:
        qs = qs.filter(
            Q(fornecedor__nome__icontains=busca) | Q(observacao__icontains=busca)
        )
    if status_selecionado:
        qs = qs.filter(status=status_selecionado)
    if fornecedor_selecionado == 'SEM_FORNECEDOR':
        qs = qs.filter(fornecedor__isnull=True)
    elif fornecedor_selecionado:
        qs = qs.filter(fornecedor__nome=fornecedor_selecionado)

    fornecedores_unicos = list(
        Fornecedor.objects.filter(ordemcompra__isnull=False)
        .values_list('nome', flat=True).distinct().order_by('nome')
    )

    paginator = Paginator(qs, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    extra_params = f"&busca={busca}&status={status_selecionado}&fornecedor={fornecedor_selecionado}"
    tem_filtros_ativos = bool(busca or status_selecionado or fornecedor_selecionado)

    context = {
        'page_obj': page_obj,
        'ordens': page_obj,
        'busca': busca,
        'status_selecionado': status_selecionado,
        'fornecedor_selecionado': fornecedor_selecionado,
        'fornecedores_unicos': fornecedores_unicos,
        'status_choices': OrdemCompra.STATUS_CHOICES,
        'extra_params': extra_params,
        'tem_filtros_ativos': tem_filtros_ativos,
    }

    if request.headers.get('HX-Request'):
        return render(request, 'estoque/ordens/_lista_resultados.html', context)
    return render(request, 'estoque/lista_ordens.html', context)


@login_required
def criar_ordem(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        with transaction.atomic():
            ordem = OrdemCompra.objects.create(
                fornecedor_id=data.get('fornecedor_id') or None,
                observacao=data.get('observacao', ''),
            )
            for item in data.get('itens', []):
                ItemOrdemCompra.objects.create(
                    ordem=ordem,
                    produto_id=item['produto_id'],
                    quantidade=item['quantidade'],
                    preco_unitario=item['preco_unitario'],
                )
        log_acao(request.user, 'CRIAR', f'Criou ordem de compra #{ordem.id}', 'OrdemCompra', ordem.id)
        return JsonResponse({'ok': True, 'id': ordem.id})
    fornecedores = Fornecedor.objects.all().values('id', 'nome')
    produtos = Produto.objects.all().values('id', 'descricao', 'preco_custo')
    return render(request, 'estoque/criar_ordem.html', {
        'fornecedores': list(fornecedores),
        'produtos': list(produtos),
    })


@login_required
def detalhe_ordem(request, id):
    ordem = get_object_or_404(OrdemCompra.objects.select_related('fornecedor'), id=id)
    itens = ordem.itens.select_related('produto').all()
    itens_total = sum(item.quantidade * item.preco_unitario for item in itens)
    return render(request, 'estoque/detalhe_ordem.html', {
        'ordem': ordem,
        'itens': itens,
        'itens_total': itens_total,
    })


@login_required
def aprovar_ordem(request, id):
    ordem = get_object_or_404(OrdemCompra, id=id)
    if ordem.status != 'PENDENTE':
        return JsonResponse({'ok': False, 'erro': 'Ordem nao esta pendente.'}, status=400)
    ordem.status = 'APROVADA'
    ordem.save()
    log_acao(request.user, 'APROVAR', f'Aprovou ordem de compra #{ordem.id}', 'OrdemCompra', id)
    return JsonResponse({'ok': True})


@login_required
def cancelar_ordem(request, id):
    ordem = get_object_or_404(OrdemCompra, id=id)
    if ordem.status in ('RECEBIDA', 'CANCELADA'):
        return JsonResponse({'ok': False, 'erro': 'Ordem ja finalizada.'}, status=400)
    ordem.status = 'CANCELADA'
    ordem.save()
    log_acao(request.user, 'CANCELAR', f'Cancelou ordem de compra #{ordem.id}', 'OrdemCompra', id)
    return JsonResponse({'ok': True})


@login_required
def receber_ordem(request, id):
    ordem = get_object_or_404(OrdemCompra.objects.select_related('fornecedor'), id=id)
    if ordem.status != 'APROVADA':
        return JsonResponse({'ok': False, 'erro': 'Ordem precisa estar aprovada para ser recebida.'}, status=400)
    with transaction.atomic():
        itens = ordem.itens.select_related('produto').all()
        for item in itens:
            produto = Produto.objects.select_for_update().get(pk=item.produto.pk)
            produto.preco_custo = item.preco_unitario
            produto.save()
            Movimentacao.objects.create(
                produto=produto,
                usuario=request.user,
                tipo='ENTRADA',
                quantidade=item.quantidade,
                observacao=f'Recebimento da Ordem #{ordem.id}',
            )
        ordem.status = 'RECEBIDA'
        ordem.save()
    log_acao(request.user, 'RECEBER', f'Recebeu ordem de compra #{ordem.id} no estoque', 'OrdemCompra', id)
    return JsonResponse({'ok': True})


@login_required
def etiqueta_produto(request, id):
    produto = get_object_or_404(Produto, id=id)
    return render(request, 'estoque/etiqueta.html', {'produto': produto})


@login_required
def relatorio_mensal(request):
    hoje = timezone.now()
    data_inicio = request.GET.get('data_inicio', hoje.replace(day=1).strftime('%Y-%m-%d'))
    data_fim = request.GET.get('data_fim', hoje.strftime('%Y-%m-%d'))
    try:
        dt_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').replace(tzinfo=timezone.get_current_timezone())
        dt_fim = datetime.strptime(data_fim, '%Y-%m-%d').replace(hour=23, minute=59, second=59, tzinfo=timezone.get_current_timezone())
    except ValueError:
        dt_inicio = hoje.replace(day=1)
        dt_fim = hoje

    movs = Movimentacao.objects.filter(data__gte=dt_inicio, data__lte=dt_fim).select_related('produto')
    total_entradas_unidade = serializar_totais_unidade(agrupar_quantidade_por_unidade(movs.filter(tipo='ENTRADA')))
    total_saidas_unidade = serializar_totais_unidade(agrupar_quantidade_por_unidade(movs.filter(tipo='SAIDA')))

    por_produto = movs.values('produto__descricao', 'produto__tipo_produto', 'produto__unidade_medida', 'tipo').annotate(
        total=Sum('quantidade')
    ).order_by('produto__descricao')

    movs_por_produto = {}
    for item in por_produto:
        nome = item['produto__descricao']
        if nome not in movs_por_produto:
            fake_produto = type('ProdutoUnidade', (), {
                'tipo_produto': item['produto__tipo_produto'],
                'unidade_medida': item['produto__unidade_medida'],
            })()
            movs_por_produto[nome] = {'entradas': Decimal('0'), 'saidas': Decimal('0'), 'unidade': unidade_base_codigo(fake_produto)}
        if item['tipo'] == 'ENTRADA':
            movs_por_produto[nome]['entradas'] += item['total']
        else:
            movs_por_produto[nome]['saidas'] += item['total']

    produtos_afetados = [
        {
            'nome': nome,
            'entradas': formatar_quantidade(d['entradas'], d['unidade']),
            'saidas': formatar_quantidade(d['saidas'], d['unidade']),
            'saldo': formatar_quantidade(d['entradas'] - d['saidas'], d['unidade']),
        }
        for nome, d in movs_por_produto.items()
    ]

    return render(request, 'estoque/relatorio.html', {
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'total_entradas_unidade': total_entradas_unidade,
        'total_saidas_unidade': total_saidas_unidade,
        'total_movimentacoes': movs.count(),
        'produtos_afetados': produtos_afetados,
    })


@login_required
def log_acoes(request):
    busca = (request.GET.get('busca') or request.GET.get('q') or '').strip()
    acao = request.GET.get('acao', '').strip()
    logs_qs = LogAcao.objects.select_related('usuario').all().order_by('-data')
    if busca:
        logs_qs = logs_qs.filter(
            Q(descricao__icontains=busca)
            | Q(usuario__username__icontains=busca)
            | Q(modelo__icontains=busca)
        )
    if acao:
        logs_qs = logs_qs.filter(acao=acao)

    paginator = Paginator(logs_qs, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    extra_params = f"&busca={busca}&acao={acao}"
    tem_filtros_ativos = bool(busca or acao)

    context = {
        'page_obj': page_obj,
        'logs': page_obj,
        'busca': busca,
        'acao_selecionada': acao,
        'acoes': LogAcao.ACAO_CHOICES,
        'extra_params': extra_params,
        'tem_filtros_ativos': tem_filtros_ativos,
    }

    if request.headers.get('HX-Request'):
        return render(request, 'estoque/logs/_lista_resultados.html', context)
    return render(request, 'estoque/log_acoes.html', context)


@login_required
def lista_usuarios(request):
    from django.contrib.auth.models import User, Group
    if not request.user.is_superuser:
        if request.method == 'POST':
            return json_erro('Permissão negada.', status=403, codigo='PERMISSAO_NEGADA')
        return render(request, 'estoque/lista_usuarios.html', {'erro': 'Apenas administradores podem gerenciar usuarios.'})
    usuarios = User.objects.prefetch_related('groups').all()
    grupos = Group.objects.all()
    perfis = [
        {'id': g.id, 'nome': PERFIS_NEGOCIO.get(g.name, g.name), 'interno': g.name}
        for g in grupos
    ]
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        acao = data.get('acao')
        if acao == 'criar':
            try:
                username = validate_username_available(data.get('username'))
                user = User.objects.create_user(
                    username=username,
                    password=data['password'],
                    is_staff=True,
                )
            except ValidationError as e:
                return JsonResponse({'ok': False, 'erro': '; '.join(e.messages)}, status=400)
            log_acao(request.user, 'CRIAR', f'Criou usuario {user.username}', 'User', user.id)
        elif acao == 'grupo':
            perm_error = exigir_admin_json(request)
            if perm_error:
                return perm_error
            user = User.objects.get(id=data['user_id'])
            grupo_id = data.get('grupo_id')
            grupo = Group.objects.get(id=grupo_id) if grupo_id else None
            novo_superuser = bool(grupo and grupo.name == 'Admin')
            admins_ativos = User.objects.filter(is_superuser=True, is_active=True).count()
            removendo_admin = user.is_superuser and not novo_superuser
            if removendo_admin and user.id == request.user.id:
                return json_erro('Você não pode remover sua própria permissão administrativa.', status=400, codigo='PERMISSAO_NEGADA')
            if removendo_admin and admins_ativos <= 1:
                return json_erro('O sistema precisa manter pelo menos um administrador.', status=400, codigo='PERMISSAO_NEGADA')
            user.groups.clear()
            if grupo:
                user.groups.add(grupo)
            user.is_staff = grupo is not None
            user.is_superuser = novo_superuser
            user.save()
            perfil_nome = PERFIS_NEGOCIO.get(grupo.name, grupo.name) if grupo else 'Sem perfil'
            log_acao(request.user, 'EDITAR', f'Alterou perfil do usuário {user.username} para {perfil_nome}', 'User', user.id)
            return json_ok(mensagem='Perfil atualizado com sucesso.')
        return json_ok()
    return render(request, 'estoque/lista_usuarios.html', {
        'usuarios': usuarios,
        'grupos': grupos,
        'perfis': perfis,
    })


@login_required
def template_csv_produtos(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="modelo_importacao_produtos.csv"'
    writer = csv.writer(response)
    writer.writerow(['descricao', 'tipo_produto', 'qt_rolos', 'metros_por_rolo', 'quantidade_base', 'qt_vidros', 'litros_por_vidro', 'preco_custo', 'preco_venda', 'estoque_minimo', 'observacao'])
    writer.writerow(['PAPEL TUCANO', 'PAPEL', '11', '500', '', '', '', '0', '0', '0', ''])
    writer.writerow(['TACTEL - ALEXANDRE', 'TECIDO', '', '', '600', '', '', '0', '0', '0', ''])
    writer.writerow(['BLACK SUBLIMACAO', 'TINTA', '', '', '', '7', '1', '0', '0', '0', ''])
    writer.writerow(['FIO NAUTICO BRANCO', 'AVIAMENTO', '', '', '19', '', '', '0', '0', '0', 'Largura 5mm'])
    return response


@login_required
def importar_csv_produtos(request):
    if request.method == 'POST' and request.FILES.get('arquivo'):
        arquivo = request.FILES['arquivo']
        if not arquivo.name.endswith('.csv'):
            return JsonResponse({'ok': False, 'erro': 'Por favor, envie um arquivo .csv válido.'}, status=400)
            
        try:
            decoded_file = arquivo.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            produtos_criados = 0
            
            with transaction.atomic():
                for row in reader:
                    keys = [k for k in row.keys() if k]
                    descricao_key = next((k for k in keys if 'descricao' in k.lower()), None)
                    if not descricao_key:
                        raise ValueError("Coluna 'descricao' não encontrada no cabeçalho.")
                    
                    descricao = row.get(descricao_key, '').strip()
                    if not descricao:
                        continue
                        
                    tipo = row.get('tipo_produto', 'OUTRO').strip().upper()
                    if tipo not in [c[0] for c in Produto.TIPO_PRODUTO_CHOICES]:
                        tipo = 'OUTRO'
                        
                    def parse_decimal(val):
                        if val in (None, ''):
                            return None
                        text = str(val).replace(',', '.').strip()
                        return Decimal(text) if text else None

                    def parse_decimal_zero(val):
                        parsed = parse_decimal(val)
                        return parsed if parsed is not None else Decimal('0')
                        
                    qt_rolos = parse_decimal_zero(row.get('qt_rolos'))
                    metros_por_rolo = parse_decimal_zero(row.get('metros_por_rolo'))
                    qt_vidros = parse_decimal_zero(row.get('qt_vidros'))
                    litros_por_vidro = parse_decimal_zero(row.get('litros_por_vidro'))
                    quantidade_base = parse_decimal_zero(row.get('quantidade_base'))
                    
                    if tipo in ['TECIDO', 'PAPEL'] and qt_rolos > 0 and metros_por_rolo > 0 and quantidade_base == 0:
                        quantidade_base = qt_rolos * metros_por_rolo
                    elif tipo == 'TINTA' and qt_vidros > 0 and litros_por_vidro > 0 and quantidade_base == 0:
                        quantidade_base = qt_vidros * litros_por_vidro
                        
                    desc_com_obs = descricao
                    obs = row.get('observacao', '').strip()
                    if tipo == 'AVIAMENTO' and obs:
                        desc_com_obs = f"{descricao} ({obs})"
                        
                    Produto.objects.create(
                        descricao=desc_com_obs,
                        tipo_produto=tipo,
                        quantidade_base=quantidade_base,
                        metros_por_rolo=metros_por_rolo if metros_por_rolo > 0 else None,
                        litros_por_vidro=litros_por_vidro if litros_por_vidro > 0 else None,
                        preco_custo=parse_decimal(row.get('preco_custo')),
                        preco_venda=parse_decimal(row.get('preco_venda')),
                        estoque_minimo=parse_decimal(row.get('estoque_minimo')),
                    )
                    produtos_criados += 1
            
            log_acao(request.user, 'CRIAR', f'Importou {produtos_criados} produtos via CSV', 'Produto')
            return JsonResponse({'ok': True, 'mensagem': f'{produtos_criados} produtos importados com sucesso!'})
            
        except Exception as e:
            return JsonResponse({'ok': False, 'erro': f'Erro ao processar arquivo: {str(e)}'}, status=400)
            
    return JsonResponse({'ok': False, 'erro': 'Método não permitido ou arquivo não enviado.'}, status=400)


# ─────────────────────────────────────────────
#  FORNECEDORES
# ─────────────────────────────────────────────

@login_required
def lista_fornecedores(request):
    busca = request.GET.get('q', '').strip()
    contexto = contexto_lista_fornecedores(busca)
    if requisicao_htmx(request) and request.headers.get('HX-Target') == 'fornecedores-resultados':
        return render(request, 'estoque/fornecedores/_resultados.html', contexto)
    return render(request, 'estoque/fornecedores/lista.html', contexto)


def contexto_lista_fornecedores(busca=''):
    fornecedores = Fornecedor.objects.annotate(total_produtos=Count('produto')).order_by('nome')
    total_fornecedores = fornecedores.count()
    if busca:
        fornecedores = fornecedores.filter(
            Q(nome__icontains=busca)
            | Q(cnpj__icontains=busca)
            | Q(email__icontains=busca)
            | Q(telefone__icontains=busca)
        )
    return {
        'fornecedores': fornecedores,
        'total_fornecedores': total_fornecedores,
        'busca': busca,
    }


def resposta_erro_fornecedor(request, mensagem, status=400):
    if requisicao_htmx(request):
        response = render(
            request,
            'estoque/fornecedores/_form_feedback.html',
            {'erro': mensagem},
        )
        response['HX-Retarget'] = '#fornecedor-form-feedback'
        response['HX-Reswap'] = 'innerHTML'
        return response
    return json_erro(mensagem, status=status)


@login_required
def salvar_fornecedor(request, id=None):
    """Cria ou edita um fornecedor via formulário HTMX ou JSON."""
    fornecedor = get_object_or_404(Fornecedor, id=id) if id else None
    if request.method == 'GET':
        return render(request, 'estoque/fornecedores/_form_modal.html', {
            'fornecedor': fornecedor,
            'busca': request.GET.get('q', '').strip(),
        })
    if request.method != 'POST':
        return json_erro('Método não permitido.', status=405)

    try:
        data = request.POST if requisicao_htmx(request) else json.loads(request.body)
    except json.JSONDecodeError:
        return resposta_erro_fornecedor(request, 'JSON inválido.')

    nome = data.get('nome', '').strip()
    if not nome:
        return resposta_erro_fornecedor(request, 'Nome é obrigatório.')

    acao = 'EDITAR' if fornecedor else 'CRIAR'
    fornecedor = fornecedor or Fornecedor()

    fornecedor.nome = nome
    fornecedor.cnpj = (data.get('cnpj') or '').strip()
    fornecedor.email = (data.get('email') or '').strip()
    fornecedor.telefone = (data.get('telefone') or '').strip()
    fornecedor.observacao = (data.get('observacao') or '').strip()
    try:
        fornecedor.full_clean()
        fornecedor.save()
    except ValidationError as exc:
        return resposta_erro_fornecedor(request, '; '.join(exc.messages))

    log_acao(request.user, acao, f'{acao} fornecedor: {fornecedor.nome}', 'Fornecedor', fornecedor.id)
    if requisicao_htmx(request):
        response = render(
            request,
            'estoque/fornecedores/_resultados.html',
            contexto_lista_fornecedores(data.get('q', '').strip()),
        )
        response['HX-Trigger-After-Swap'] = json.dumps({
            'sge:feedback': {
                'message': 'Fornecedor atualizado com sucesso.' if id else 'Fornecedor criado com sucesso.',
                'type': 'success',
            },
            'sge:modal-close': {'id': 'modalFornecedor'},
        })
        return response
    return json_ok(id=fornecedor.id)


@login_required
def excluir_fornecedor(request, id):
    if request.method != 'POST':
        return json_erro('Exclusão deve usar POST.', status=405)
    perm_error = exigir_admin_json(request)
    if perm_error:
        return perm_error
    fornecedor = get_object_or_404(Fornecedor, id=id)
    vinculados = Produto.objects.filter(fornecedor=fornecedor).count()
    if vinculados:
        mensagem = f'O fornecedor não pode ser excluído porque possui {vinculados} produto(s) vinculado(s).'
        if requisicao_htmx(request):
            response = HttpResponse()
            response['HX-Reswap'] = 'none'
            response['HX-Trigger'] = json.dumps({
                'sge:feedback': {'message': mensagem, 'type': 'danger'},
            })
            return response
        return json_erro(
            mensagem,
            codigo='VINCULO_IMPEDITIVO',
        )
    nome = fornecedor.nome
    fornecedor.delete()
    log_acao(request.user, 'EXCLUIR', f'Excluiu fornecedor: {nome}', 'Fornecedor', id)
    if requisicao_htmx(request):
        response = render(
            request,
            'estoque/fornecedores/_resultados.html',
            contexto_lista_fornecedores(request.POST.get('q', '').strip()),
        )
        response['HX-Trigger-After-Swap'] = json.dumps({
            'sge:feedback': {'message': 'Fornecedor excluído com sucesso.', 'type': 'success'},
        })
        return response
    return json_ok(mensagem='Fornecedor excluído com sucesso.')


# ─────────────────────────────────────────────
#  CATEGORIAS
# ─────────────────────────────────────────────

@login_required
def lista_categorias(request):
    busca = request.GET.get('q', '').strip()
    contexto = contexto_lista_categorias(busca)
    if requisicao_htmx(request) and request.headers.get('HX-Target') == 'categorias-resultados':
        return render(request, 'estoque/categorias/_resultados.html', contexto)
    return render(request, 'estoque/categorias/lista.html', contexto)


def contexto_lista_categorias(busca=''):
    categorias = Categoria.objects.annotate(total_produtos=Count('produtos')).order_by('nome')
    total_categorias = categorias.count()
    if busca:
        categorias = categorias.filter(
            Q(nome__icontains=busca) | Q(descricao__icontains=busca)
        )
    return {
        'categorias': categorias,
        'total_categorias': total_categorias,
        'busca': busca,
    }


def resposta_erro_categoria(request, mensagem, status=400):
    if requisicao_htmx(request):
        response = render(
            request,
            'estoque/categorias/_form_feedback.html',
            {'erro': mensagem},
        )
        response['HX-Retarget'] = '#categoria-form-feedback'
        response['HX-Reswap'] = 'innerHTML'
        return response
    return json_erro(mensagem, status=status)


@login_required
def salvar_categoria(request, id=None):
    """Cria ou edita uma categoria via formulário HTMX ou JSON."""
    categoria = get_object_or_404(Categoria, id=id) if id else None
    if request.method == 'GET':
        return render(request, 'estoque/categorias/_form_modal.html', {
            'categoria': categoria,
            'busca': request.GET.get('q', '').strip(),
        })
    if request.method != 'POST':
        return json_erro('Método não permitido.', status=405)

    try:
        data = request.POST if requisicao_htmx(request) else json.loads(request.body)
    except json.JSONDecodeError:
        return resposta_erro_categoria(request, 'JSON inválido.')

    nome = data.get('nome', '').strip()
    if not nome:
        return resposta_erro_categoria(request, 'Nome é obrigatório.')

    cor = (data.get('cor') or '#6c757d').strip()
    if len(cor) != 7 or not cor.startswith('#') or not all(
        caractere in '0123456789abcdefABCDEF' for caractere in cor[1:]
    ):
        return resposta_erro_categoria(request, 'Cor inválida. Use o formato hexadecimal #RRGGBB.')

    acao = 'EDITAR' if categoria else 'CRIAR'
    categoria = categoria or Categoria()

    categoria.nome = nome
    categoria.descricao = (data.get('descricao') or '').strip()
    categoria.cor = cor
    try:
        categoria.full_clean()
        categoria.save()
    except ValidationError as exc:
        return resposta_erro_categoria(request, '; '.join(exc.messages))
    except IntegrityError:
        return resposta_erro_categoria(request, 'Já existe uma categoria com este nome.')

    log_acao(request.user, acao, f'{acao} categoria: {categoria.nome}', 'Categoria', categoria.id)
    if requisicao_htmx(request):
        response = render(
            request,
            'estoque/categorias/_resultados.html',
            contexto_lista_categorias(data.get('q', '').strip()),
        )
        response['HX-Trigger-After-Swap'] = json.dumps({
            'sge:feedback': {
                'message': 'Categoria atualizada com sucesso.' if id else 'Categoria criada com sucesso.',
                'type': 'success',
            },
            'sge:modal-close': {'id': 'modalCategoria'},
        })
        return response
    return json_ok(id=categoria.id)


@login_required
def excluir_categoria(request, id):
    if request.method != 'POST':
        return json_erro('Exclusão deve usar POST.', status=405)
    perm_error = exigir_admin_json(request)
    if perm_error:
        return perm_error
    categoria = get_object_or_404(Categoria, id=id)
    vinculados = categoria.produtos.count()
    if vinculados:
        mensagem = f'A categoria não pode ser excluída porque possui {vinculados} produto(s) vinculado(s).'
        if requisicao_htmx(request):
            response = HttpResponse()
            response['HX-Reswap'] = 'none'
            response['HX-Trigger'] = json.dumps({
                'sge:feedback': {'message': mensagem, 'type': 'danger'},
            })
            return response
        return json_erro(
            mensagem,
            codigo='VINCULO_IMPEDITIVO',
        )
    nome = categoria.nome
    categoria.delete()
    log_acao(request.user, 'EXCLUIR', f'Excluiu categoria: {nome}', 'Categoria', id)
    if requisicao_htmx(request):
        response = render(
            request,
            'estoque/categorias/_resultados.html',
            contexto_lista_categorias(request.POST.get('q', '').strip()),
        )
        response['HX-Trigger-After-Swap'] = json.dumps({
            'sge:feedback': {'message': 'Categoria excluída com sucesso.', 'type': 'success'},
        })
        return response
    return json_ok(mensagem='Categoria excluída com sucesso.')


@login_required
def lista_fechamentos(request):
    hoje = timezone.localdate()
    data_inicio = hoje.replace(day=1)
    data_fim = hoje.replace(day=calendar.monthrange(hoje.year, hoje.month)[1])
    return render(request, 'estoque/fechamentos.html', {
        **contexto_lista_fechamentos(),
        'data_inicio_sugerida': data_inicio.isoformat(),
        'data_fim_sugerida': data_fim.isoformat(),
        'resumo_inicial': resumo_fechamento(data_inicio, data_fim),
    })


def contexto_lista_fechamentos():
    fechamentos = FechamentoMensal.objects.prefetch_related('itens').select_related('usuario').all()
    fechamentos_data = []
    for f in fechamentos:
        itens = list(f.itens.all())
        total_itens = len(itens)
        valor_total = Decimal('0.00')
        produtos_sem_custo = 0
        for item in itens:
            if item.quantidade > 0 and item.preco_custo is None:
                produtos_sem_custo += 1
            elif item.preco_custo is not None:
                valor_total += item.quantidade * item.preco_custo
        fechamentos_data.append({
            'id': f.id,
            'data_fechamento': f.data_fechamento.strftime('%d/%m/%Y %H:%M'),
            'usuario': f.usuario.username if f.usuario else '-',
            'periodo_formatado': f.periodo_formatado,
            'observacao': f.observacao,
            'total_itens': total_itens,
            'valor_total_formatado': dinheiro_br(valor_total),
            'produtos_sem_custo': produtos_sem_custo,
            'calculo_completo': produtos_sem_custo == 0,
        })
    return {'fechamentos': fechamentos_data}


@login_required
def detalhe_fechamento(request, id):
    fechamento = get_object_or_404(
        FechamentoMensal.objects.select_related('usuario'),
        id=id,
    )
    itens = []
    valor_total = Decimal('0.00')
    valor_total_venda = Decimal('0.00')
    produtos_sem_custo = 0
    for item in fechamento.itens.all().order_by('descricao'):
        valor_custo = None
        if item.preco_custo is None:
            if item.quantidade > 0:
                produtos_sem_custo += 1
        else:
            valor_custo = item.quantidade * item.preco_custo
            valor_total += valor_custo
        valor_venda = None
        if item.preco_venda is not None:
            valor_venda = item.quantidade * item.preco_venda
            valor_total_venda += valor_venda
        unidade = UNIDADES.get(item.unidade_medida, UNIDADES['OUTRO'])
        itens.append({
            'descricao': item.descricao,
            'tipo': item.get_tipo_produto_display() or '—',
            'categoria': item.categoria_nome or '—',
            'fornecedor': item.fornecedor_nome or '—',
            'unidade': unidade.simbolo or '—',
            'quantidade': item.quantidade,
            'preco_custo': item.preco_custo,
            'preco_venda': item.preco_venda,
            'valor_custo': valor_custo,
            'valor_venda': valor_venda,
        })
    return render(request, 'estoque/detalhe_fechamento.html', {
        'fechamento': fechamento,
        'itens': itens,
        'valor_total': valor_total,
        'valor_total_formatado': dinheiro_br(valor_total),
        'valor_total_venda': valor_total_venda,
        'valor_total_venda_formatado': dinheiro_br(valor_total_venda),
        'produtos_sem_custo': produtos_sem_custo,
    })


@login_required
def excluir_fechamento(request, id):
    if request.method != 'POST':
        return json_erro('Exclusão deve usar POST.', status=405)
    perm_error = exigir_admin_json(request)
    if perm_error:
        return perm_error

    with transaction.atomic():
        fechamento = get_object_or_404(
            FechamentoMensal.objects.select_for_update(),
            id=id,
        )
        periodo = fechamento.periodo_formatado
        total_itens = fechamento.itens.count()
        fechamento.delete()
        log_acao(
            request.user,
            'EXCLUIR',
            f'Excluiu fechamento de estoque de {periodo} com {total_itens} item(ns)',
            'FechamentoMensal',
            id,
        )

    if requisicao_htmx(request):
        response = render(
            request,
            'estoque/fechamentos/_lista.html',
            contexto_lista_fechamentos(),
        )
        response['HX-Trigger-After-Swap'] = json.dumps({
            'sge:feedback': {
                'message': 'Fechamento excluído. O período já pode receber um novo freeze.',
                'type': 'success',
            },
        })
        return response
    return json_ok(mensagem='Fechamento excluído. O período já pode receber um novo freeze.')


@login_required
def revisar_fechamento(request):
    try:
        inicio = data_iso(request.GET.get('data_inicio'), 'Data inicial')
        fim = data_iso(request.GET.get('data_fim'), 'Data final')
        resumo = resumo_fechamento(inicio, fim)
        if resumo['duplicado']:
            mensagem = f'Já existe um fechamento para {resumo["periodo_formatado"]}.'
            if requisicao_htmx(request):
                return render(request, 'estoque/fechamentos/_revisao.html', {'erro': mensagem})
            return json_erro(
                mensagem,
                codigo='FECHAMENTO_DUPLICADO',
            )
        if requisicao_htmx(request):
            return render(request, 'estoque/fechamentos/_revisao.html', {'resumo': resumo})
        return json_ok(resumo=resumo)
    except ValidationError as exc:
        if requisicao_htmx(request):
            return render(
                request,
                'estoque/fechamentos/_revisao.html',
                {'erro': '; '.join(exc.messages)},
            )
        return json_erro('; '.join(exc.messages))


@login_required
def realizar_fechamento(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'erro': 'Método não permitido.'}, status=405)
    try:
        data = request.POST if requisicao_htmx(request) else json.loads(request.body)
        inicio = data_iso(data.get('data_inicio'), 'Data inicial')
        fim = data_iso(data.get('data_fim'), 'Data final')
        observacao = (data.get('observacao') or '').strip()
        fechamento = criar_fechamento_periodo(
            data_inicio=inicio,
            data_fim=fim,
            usuario=request.user,
            observacao=observacao,
        )
        if requisicao_htmx(request):
            messages.success(request, 'Fechamento de estoque realizado com sucesso.')
            response = HttpResponse()
            response['HX-Redirect'] = reverse('lista_fechamentos')
            return response
        return json_ok(mensagem='Fechamento de estoque realizado com sucesso.', id=fechamento.id)
    except json.JSONDecodeError:
        return json_erro('JSON inválido.')
    except ValidationError as exc:
        mensagem = '; '.join(exc.messages)
        if requisicao_htmx(request):
            return render(request, 'estoque/fechamentos/_revisao.html', {'erro': mensagem})
        codigo = 'FECHAMENTO_DUPLICADO' if 'Já existe' in mensagem else 'VALIDACAO'
        return json_erro(mensagem, codigo=codigo)


@login_required
def exportar_fechamento_xlsx(request, id):
    fechamento = get_object_or_404(FechamentoMensal, id=id)
    itens = fechamento.itens.all().order_by('descricao')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Fechamento {fechamento.data_fim:%Y-%m-%d}"

    font_title = Font(name='Segoe UI', size=16, bold=True, color='1E293B')
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_data = Font(name='Segoe UI', size=11, color='1E293B')
    font_total = Font(name='Segoe UI', size=11, bold=True, color='1E293B')
    
    fill_header = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    fill_total = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    ws.merge_cells('A1:J1')
    ws['A1'] = f"S.G.E - Fechamento de Estoque ({fechamento.periodo_formatado})"
    ws['A1'].font = font_title
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Realizado em: {fechamento.data_fechamento.strftime('%d/%m/%Y %H:%M')} por {fechamento.usuario.username if fechamento.usuario else '-'}"
    ws['A2'].font = Font(name='Segoe UI', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    headers = [
        "Descrição do Material", "Tipo", "Categoria", "Fornecedor",
        "Unid.", "Quantidade", "Preço Custo",
        "Preço Venda", "Total Custo", "Total Venda"
    ]
    
    ws.append([])
    ws.row_dimensions[4].height = 28
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center' if col_num > 4 else 'left', vertical='center')
        cell.border = border_thin

    start_row = 5
    for item in itens:
        tipo = item.get_tipo_produto_display() or '-'
        fornecedor = item.fornecedor_nome or '-'
        unidade_info_snapshot = UNIDADES.get(item.unidade_medida, UNIDADES['OUTRO'])
        unidade = unidade_info_snapshot.simbolo
        preco_custo = item.preco_custo
        preco_venda = item.preco_venda
        next_row = ws.max_row + 1
        
        row_data = [
            item.descricao,
            tipo,
            item.categoria_nome or '-',
            fornecedor,
            unidade,
            float(item.quantidade),
            float(preco_custo) if preco_custo is not None else None,
            float(preco_venda) if preco_venda is not None else None,
            f"=F{next_row}*G{next_row}" if preco_custo is not None else None,
            f"=F{next_row}*H{next_row}" if preco_venda is not None else None,
        ]
        
        ws.append(row_data)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 20
        
        for col_idx in range(1, 11):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = font_data
            cell.border = border_thin
            
            if col_idx in (5, 6):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in (7, 8, 9, 10):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
            if col_idx in (7, 8, 9, 10):
                cell.number_format = 'R$ #,##0.00'
            elif col_idx == 6:
                cell.number_format = '#,##0.00'

    end_row = ws.max_row
    ws.append([
        "TOTAL GERAL", "", "", "", "",
        "Quantidades por unidade não são somadas", "", "",
        f"=SUM(I{start_row}:I{end_row})",
        f"=SUM(J{start_row}:J{end_row})"
    ])
    
    total_row = ws.max_row
    ws.row_dimensions[total_row].height = 26
    
    for col_idx in range(1, 11):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = font_total
        cell.fill = fill_total
        cell.border = border_thin
        
        if col_idx in (6, 9, 10):
            cell.alignment = Alignment(horizontal='left' if col_idx == 6 else 'right', vertical='center')
            if col_idx in (9, 10):
                cell.number_format = 'R$ #,##0.00'
            else:
                cell.number_format = '#,##0.00'
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.row > 2 and cell.value:
                val_str = str(cell.value)
                if val_str.startswith('='):
                    val_str = "R$ 999.999,99"
                max_len = max(max_len, len(val_str))
                
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename="fechamento_{fechamento.data_inicio:%d-%m-%Y}'
        f'_a_{fechamento.data_fim:%d-%m-%Y}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def exportar_atual_xlsx(request):
    produtos = Produto.objects.select_related('fornecedor').all().order_by('descricao')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque Atual"

    font_title = Font(name='Segoe UI', size=16, bold=True, color='1E293B')
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_data = Font(name='Segoe UI', size=11, color='1E293B')
    font_total = Font(name='Segoe UI', size=11, bold=True, color='1E293B')
    
    fill_header = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    fill_total = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    ws.merge_cells('A1:I1')
    ws['A1'] = "S.G.E - Relatório de Posição de Estoque Atual"
    ws['A1'].font = font_title
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:I2')
    ws['A2'] = f"Gerado em: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(name='Segoe UI', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    headers = [
        "Descrição do Material", "Tipo", "Fornecedor", 
        "Unid.", "Quantidade", "Preço Custo", 
        "Preço Venda", "Total Custo", "Total Venda"
    ]
    
    ws.append([])
    ws.row_dimensions[4].height = 28
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center' if col_num > 3 else 'left', vertical='center')
        cell.border = border_thin

    start_row = 5
    for p in produtos:
        tipo = p.get_tipo_produto_display()
        fornecedor = p.fornecedor.nome if p.fornecedor else '-'
        unidade = p.unidade_simbolo
        preco_custo = p.preco_custo
        preco_venda = p.preco_venda
        
        row_data = [
            p.descricao,
            tipo,
            fornecedor,
            unidade,
            float(p.quantidade_base),
            float(preco_custo) if preco_custo is not None else None,
            float(preco_venda) if preco_venda is not None else None,
            f"=E{ws.max_row+1}*F{ws.max_row+1}" if preco_custo is not None else None,
            f"=E{ws.max_row+1}*G{ws.max_row+1}" if preco_venda is not None else None,
        ]
        
        ws.append(row_data)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 20
        
        for col_idx in range(1, 10):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = font_data
            cell.border = border_thin
            
            if col_idx in (4, 5):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in (6, 7, 8, 9):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
            if col_idx in (6, 7, 8, 9):
                cell.number_format = 'R$ #,##0.00'
            elif col_idx == 5:
                cell.number_format = '#,##0.00'

    end_row = ws.max_row
    ws.append([
        "TOTAL GERAL", "", "", "", 
        "Quantidades por unidade não são somadas", "", "",
        f"=SUM(H{start_row}:H{end_row})", 
        f"=SUM(I{start_row}:I{end_row})"
    ])
    
    total_row = ws.max_row
    ws.row_dimensions[total_row].height = 26
    
    for col_idx in range(1, 10):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = font_total
        cell.fill = fill_total
        cell.border = border_thin
        
        if col_idx in (5, 8, 9):
            cell.alignment = Alignment(horizontal='left' if col_idx == 5 else 'right', vertical='center')
            if col_idx in (8, 9):
                cell.number_format = 'R$ #,##0.00'
            else:
                cell.number_format = '#,##0.00'
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 2 and cell.value:
                val_str = str(cell.value)
                if val_str.startswith('='):
                    val_str = "R$ 999.999,99"
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_posicao_estoque_atual.xlsx"'
    wb.save(response)
    return response


@login_required
def busca_rapida(request):
    q = request.GET.get('q', '').strip()
    if not q:
        return JsonResponse({'resultados': []})
    
    produtos = Produto.objects.filter(
        Q(descricao__icontains=q) | Q(fornecedor__nome__icontains=q)
    ).select_related('fornecedor')[:10]
    
    resultados = [
        {
            'id': p.id,
            'descricao': p.descricao,
            'tipo_produto': p.get_tipo_produto_display(),
            'quantidade': p.quantidade_formatada,
            'unidade': p.unidade_simbolo,
        }
        for p in produtos
    ]
    return JsonResponse({'resultados': resultados})


# ─────────────────────────────────────────────────────────────────────────────
#  INTEGRAÇÃO OMIE — Importação de Notas de Entrada
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def buscar_notas_omie(request):
    """
    Exibe a lista de Notas de Entrada do Omie e permite ao usuário
    importá-las como movimentações de ENTRADA no estoque.

    GET: Lista notas do Omie (com paginação e busca por fornecedor/data), marcando quais já foram importadas.
    """
    from .models import ConfiguracaoOmie, ImportacaoNFe
    from .services.omie_client import OmieClient, OmieAPIError, OmieConfigError

    pagina = int(request.GET.get('pagina', 1))
    busca = request.GET.get('q', '').strip()
    cnpj_fornecedor = request.GET.get('cnpj', '').strip()
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()

    erro = None
    notas = []
    total_paginas = 1
    total_registros = 0
    ja_importados = set()
    config_omie = ConfiguracaoOmie.objects.first()

    try:
        client = OmieClient()
        notas, total_paginas, total_registros = client.listar_notas_parseadas(
            pagina=pagina,
            registros_por_pagina=20,
            cnpj_fornecedor=cnpj_fornecedor,
            data_inicio=data_inicio,
            data_fim=data_fim,
            ordenar_decrescente=True,
        )

        # Se o usuário digitou algo no campo de busca 'q', faz o filtro complementar
        if busca:
            q_lower = busca.lower()
            notas = [
                n for n in notas
                if q_lower in n.fornecedor_nome.lower()
                or q_lower in n.fornecedor_cnpj.lower()
                or q_lower in n.numero_nfe.lower()
                or any(q_lower in item.descricao.lower() for item in n.itens)
            ]

        # Identificar notas já importadas
        ids_notas = {n.n_cod_nota_ent for n in notas}
        ja_importados = set(
            ImportacaoNFe.objects.filter(
                n_cod_nota_ent__in=ids_notas
            ).values_list('n_cod_nota_ent', flat=True)
        )
    except OmieConfigError as exc:
        erro = f'Credenciais Omie não configuradas: {exc}'
    except OmieAPIError as exc:
        erro = f'Erro na API Omie [{exc.codigo}]: {exc.descricao}'
    except Exception as exc:
        erro = f'Erro inesperado ao conectar ao Omie: {exc}'

    # Montar lista de produtos SGE para o dropdown de seleção manual
    produtos_sge = list(
        Produto.objects.order_by('descricao').values('id', 'descricao', 'unidade_medida')
    )

    return render(request, 'estoque/omie_notas.html', {
        'notas': notas,
        'ja_importados': ja_importados,
        'erro': erro,
        'pagina': pagina,
        'busca': busca,
        'cnpj_fornecedor': cnpj_fornecedor,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'total_paginas': total_paginas,
        'total_registros': total_registros,
        'produtos_sge_json': json.dumps(produtos_sge),
        'config_omie': config_omie,
    })


@login_required
def salvar_configuracao_omie(request):
    """
    POST: Salva ou atualiza a App Key e o App Secret do Omie no banco de dados (ConfiguracaoOmie).
    Apenas administradores podem alterar.
    """
    from .models import ConfiguracaoOmie

    if request.method != 'POST':
        return json_erro('Método não permitido.', status=405)

    perm_error = exigir_admin_json(request)
    if perm_error:
        return perm_error

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return json_erro('JSON inválido.')

    app_key = data.get('app_key', '').strip().strip('"\':')
    app_secret = data.get('app_secret', '').strip().strip('"\':')

    if not app_key or not app_secret:
        return json_erro('App Key e App Secret são obrigatórios.')

    # Tratar digitação acidental da letra 'O' no lugar do número '0' em chaves hexadecimais do Omie
    if len(app_secret) == 32 and app_secret[0] in ('O', 'o'):
        import re
        if re.match(r'^[0-9a-fA-F]{31}$', app_secret[1:]):
            app_secret = '0' + app_secret[1:]

    config, _ = ConfiguracaoOmie.objects.get_or_create(id=1)
    config.app_key = app_key
    config.app_secret = app_secret
    config.usuario = request.user
    config.save()

    log_acao(
        request.user,
        'EDITAR',
        'Atualizou credenciais de API do Omie (App Key e App Secret)',
        'ConfiguracaoOmie',
        config.id,
    )

    return json_ok(mensagem='Credenciais do Omie salvas com sucesso!')


@login_required
def importar_nota_omie(request, n_cod: int):
    """
    POST: Importa uma nota de entrada do Omie gerando Movimentacoes de ENTRADA.

    Body JSON esperado:
    {
      "itens": [
        {
          "cod_item_int": "IT...",
          "produto_id": 42,        // ID do produto no SGE
          "quantidade": "10.5",
          "descricao": "Nome no Omie",
          "valor_unitario": "12.50"
        },
        ...
      ],
      "fornecedor_nome": "Fornecedor X",
      "numero_nfe": "123456",
      "cod_int_nota_ent": "NE..."
    }
    """
    from .models import ImportacaoNFe
    from .services.omie_client import OmieConfigError

    if request.method != 'POST':
        return json_erro('Método não permitido.', status=405)

    # Verificar idempotência
    if ImportacaoNFe.objects.filter(n_cod_nota_ent=n_cod).exists():
        return json_erro(
            f'A nota Omie #{n_cod} já foi importada anteriormente.',
            codigo='JA_IMPORTADO',
        )

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return json_erro('JSON inválido.')

    itens = data.get('itens', [])
    if not itens:
        return json_erro('Nenhum item para importar.')

    # Validar que todos os itens têm produto_id selecionado ou flag de criar novo
    for item in itens:
        pid = str(item.get('produto_id', ''))
        if not pid or (pid != 'novo' and not item.get('criar_novo') and not pid.isdigit()):
            descricao = item.get('descricao', '?')
            return json_erro(
                f'O item "{descricao}" não tem produto do SGE selecionado.',
                codigo='PRODUTO_NAO_SELECIONADO',
            )

    movimentacoes_criadas = 0
    descricoes_importadas = []

    try:
        with transaction.atomic():
            for item in itens:
                pid = str(item.get('produto_id', ''))
                valor_unitario = item.get('valor_unitario', '0')

                if pid == 'novo' or item.get('criar_novo'):
                    nova_desc = (item.get('novo_descricao') or item.get('descricao') or 'Novo Produto').strip()
                    tipo_prod = item.get('novo_tipo_produto', 'OUTRO')
                    unid = item.get('novo_unidade_medida', 'UN')
                    val_unit_dec = Decimal(str(valor_unitario)) if valor_unitario and Decimal(str(valor_unitario)) > 0 else Decimal('0.00')
                    est_min_raw = item.get('novo_estoque_minimo', '0')
                    est_min_dec = Decimal(str(est_min_raw)) if est_min_raw else Decimal('0.00')

                    produto = Produto.objects.create(
                        descricao=nova_desc,
                        tipo_produto=tipo_prod,
                        unidade_medida=unid,
                        quantidade_base=Decimal('0.00'),
                        preco_custo=val_unit_dec,
                        preco_venda=Decimal('0.00'),
                        estoque_minimo=est_min_dec,
                    )
                    log_acao(
                        request.user,
                        'CRIAR',
                        f'Produto "{produto.descricao}" cadastrado via importação NF-e Omie #{n_cod}',
                        'Produto',
                        produto.id,
                    )
                else:
                    produto = Produto.objects.select_for_update().get(pk=item['produto_id'])

                quantidade = Decimal(str(item.get('quantidade', '0')))
                if quantidade <= 0:
                    raise ValidationError(f'Quantidade inválida para "{produto.descricao}".')

                obs = (
                    f'Importado da NF-e Omie #{n_cod} | '
                    f'{item.get("descricao", "")} | '
                    f'Qtde: {quantidade} | '
                    f'Valor unit.: R$ {valor_unitario}'
                )

                Movimentacao.objects.create(
                    produto=produto,
                    usuario=request.user,
                    tipo='ENTRADA',
                    quantidade=quantidade,
                    observacao=obs[:255],
                )
                # Atualizar preço de custo se produto existente e preço informado
                if str(item.get('produto_id', '')) != 'novo' and not item.get('criar_novo'):
                    if valor_unitario and Decimal(str(valor_unitario)) > 0:
                        produto.preco_custo = Decimal(str(valor_unitario))
                        produto.save(update_fields=['preco_custo'])

                movimentacoes_criadas += 1
                descricoes_importadas.append(produto.descricao)

            # Registrar importação para idempotência
            ImportacaoNFe.objects.create(
                n_cod_nota_ent=n_cod,
                cod_int_nota_ent=data.get('cod_int_nota_ent', ''),
                numero_nfe=data.get('numero_nfe', ''),
                fornecedor_nome=data.get('fornecedor_nome', ''),
                usuario=request.user,
                observacao=f'{movimentacoes_criadas} itens importados',
            )

        # Auditoria
        log_acao(
            request.user,
            'ENTRADA',
            (
                f'Importação NF-e Omie #{n_cod} — '
                f'{movimentacoes_criadas} movimentação(ões) criada(s): '
                f'{", ".join(descricoes_importadas[:5])}'
                + ('...' if len(descricoes_importadas) > 5 else '')
            ),
            'ImportacaoNFe',
        )

        return json_ok(
            mensagem=f'{movimentacoes_criadas} entrada(s) registrada(s) com sucesso!',
            movimentacoes_criadas=movimentacoes_criadas,
        )

    except Produto.DoesNotExist:
        return json_erro('Produto SGE não encontrado.', status=404)
    except ValidationError as exc:
        return json_erro('; '.join(exc.messages), codigo='VALIDACAO')
    except (InvalidOperation, TypeError, ValueError) as exc:
        return json_erro(f'Valor inválido: {exc}', codigo='VALIDACAO')
    except OmieConfigError as exc:
        return json_erro(str(exc), codigo='CONFIG_ERROR', status=500)
    except Exception as exc:
        logger_omie = __import__('logging').getLogger(__name__)
        logger_omie.exception('Erro ao importar nota Omie #%s', n_cod)
        return json_erro(f'Erro interno: {exc}', status=500)
