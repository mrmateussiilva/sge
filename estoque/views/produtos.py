import csv
import json
from decimal import Decimal, InvalidOperation
from io import StringIO

from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..log_utils import log_acao
from ..models import Categoria, Fornecedor, HistoricoPreco, ItemOrdemCompra, Movimentacao, Produto
from .helpers import (
    decimal_ou_none,
    exigir_admin_json,
    json_erro,
    json_ok,
    produto_lista_vue_json,
    produto_operacional_json,
)


def registrar_ajuste_saldo(produto, usuario, nova_quantidade, observacao):
    nova_quantidade = Decimal(str(nova_quantidade))
    if nova_quantidade < 0:
        raise ValidationError('Quantidade não pode ser negativa.')
    diferenca = nova_quantidade - produto.quantidade_base
    if diferenca == 0:
        return None
    return Movimentacao.objects.create(
        produto=produto,
        usuario=usuario,
        tipo='ENTRADA' if diferenca > 0 else 'SAIDA',
        quantidade=abs(diferenca),
        observacao=observacao,
    )


@login_required
def lista_produtos(request):
    produtos = list(
        Produto.objects.select_related('fornecedor', 'categoria').order_by('descricao')
    )
    produtos_data = [produto_lista_vue_json(p) for p in produtos]
    fornecedores_unicos = list(
        Fornecedor.objects.filter(produto__isnull=False)
        .values_list('nome', flat=True).distinct().order_by('nome')
    )
    return render(request, 'estoque/lista.html', {
        'produtos_json': json.dumps(produtos_data),
        'fornecedores_json': json.dumps(fornecedores_unicos),
        'total_produtos': len(produtos_data),
    })


@login_required
def atualiza_estoque(request):
    if request.method == 'POST':
        try:
            if request.content_type == 'application/json':
                data = json.loads(request.body)
            else:
                data = request.POST

            produto_id = data.get('id')
            variacao = Decimal(str(data.get('variacao', 0)))
            produto = Produto.objects.get(id=produto_id)

            if variacao != 0:
                tipo = 'ENTRADA' if variacao > 0 else 'SAIDA'
                quantidade = abs(variacao)
                with transaction.atomic():
                    Movimentacao.objects.create(
                        produto=produto,
                        usuario=request.user,
                        tipo=tipo,
                        quantidade=quantidade,
                        observacao='Ajuste rápido de estoque',
                    )
                produto.refresh_from_db()

            if request.headers.get('HX-Request'):
                response = render(request, 'estoque/produtos/_quantidade_cell.html', {'p': produto})
                response['HX-Trigger'] = 'estoqueAtualizado'
                return response

            return json_ok(
                nova_quantidade=float(produto.quantidade_base),
                nova_quantidade_formatada=produto.quantidade_formatada,
                status_estoque=produto.status_estoque,
            )
        except (InvalidOperation, TypeError, ValueError):
            if request.headers.get('HX-Request'):
                return HttpResponse('Quantidade inválida.', status=400)
            return json_erro('Quantidade inválida.')
        except Produto.DoesNotExist:
            if request.headers.get('HX-Request'):
                return HttpResponse('Produto não encontrado.', status=404)
            return json_erro('Produto não encontrado.', status=404)
        except ValidationError as e:
            msg = '; '.join(e.messages)
            if request.headers.get('HX-Request'):
                return HttpResponse(msg, status=400)
            return json_erro(msg, codigo='SALDO_INSUFICIENTE')
    return json_erro('Método não permitido.', status=405)


@login_required
def inline_edit_estoque(request, id):
    produto = get_object_or_404(Produto, pk=id)

    if request.GET.get('cancel'):
        return render(request, 'estoque/produtos/_quantidade_cell.html', {'p': produto})

    if request.method == 'POST':
        try:
            nova_qtd = Decimal(str(request.POST.get('quantidade_base', produto.quantidade_base)))
            variacao = nova_qtd - produto.quantidade_base
            if variacao != 0:
                tipo = 'ENTRADA' if variacao > 0 else 'SAIDA'
                with transaction.atomic():
                    Movimentacao.objects.create(
                        produto=produto,
                        usuario=request.user,
                        tipo=tipo,
                        quantidade=abs(variacao),
                        observacao='Edição inline de quantidade',
                    )
                produto.refresh_from_db()
            response = render(request, 'estoque/produtos/_quantidade_cell.html', {'p': produto})
            response['HX-Trigger'] = 'estoqueAtualizado'
            return response
        except (InvalidOperation, TypeError, ValueError):
            return HttpResponse('Quantidade inválida.', status=400)
        except ValidationError as e:
            return HttpResponse('; '.join(e.messages), status=400)

    return render(request, 'estoque/produtos/_inline_edit_form.html', {'p': produto})


@login_required
def exportar_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="relatorio_estoque.csv"'

    writer = csv.writer(response)
    writer.writerow(['Descricao', 'Tipo', 'Fornecedor', 'Unidade Base', 'Quantidade', 'Preco Custo', 'Preco Venda', 'Estoque Minimo'])
    produtos = Produto.objects.select_related('fornecedor').all()
    for p in produtos:
        writer.writerow([
            p.descricao,
            p.get_tipo_produto_display(),
            p.fornecedor.nome if p.fornecedor else '',
            p.unidade_simbolo,
            p.quantidade_formatada,
            p.preco_custo if p.preco_custo is not None else '',
            p.preco_venda if p.preco_venda is not None else '',
            p.estoque_minimo if p.estoque_minimo is not None else '',
        ])
    return response


@login_required
def cadastrar_produto(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            quantidade_inicial = decimal_ou_none(data.get('quantidade_base')) or Decimal('0')
            if quantidade_inicial < 0:
                return json_erro('Quantidade não pode ser negativa.')
            with transaction.atomic():
                produto = Produto.objects.create(
                    tipo_produto=data['tipo_produto'],
                    descricao=data['descricao'],
                    fornecedor_id=data.get('fornecedor_id') or None,
                    quantidade_base=Decimal('0'),
                    preco_custo=decimal_ou_none(data.get('preco_custo')),
                    preco_venda=decimal_ou_none(data.get('preco_venda')),
                    estoque_minimo=decimal_ou_none(data.get('estoque_minimo')),
                    metros_por_rolo=data.get('metros_por_rolo') or None,
                    tipo_tinta=data.get('tipo_tinta', 'N/A'),
                    cor_tinta=data.get('cor_tinta', 'INCOLOR'),
                    litros_por_vidro=data.get('litros_por_vidro') or None,
                    unidade_medida=data.get('unidade_medida', 'UN'),
                    categoria_id=data.get('categoria_id') or None,
                )
                registrar_ajuste_saldo(
                    produto,
                    request.user,
                    quantidade_inicial,
                    'Estoque inicial do cadastro do produto',
                )
                produto.refresh_from_db()
                log_acao(request.user, 'CRIAR', f'Cadastrou produto {produto.descricao}', 'Produto', produto.id)
            return JsonResponse({'ok': True, 'id': produto.id})
        except (json.JSONDecodeError, KeyError, InvalidOperation, TypeError, ValueError):
            return json_erro('Dados inválidos.')
        except ValidationError as exc:
            return json_erro('; '.join(exc.messages))
    fornecedores = Fornecedor.objects.all().values('id', 'nome')
    categorias = Categoria.objects.all().values('id', 'nome')
    return render(request, 'estoque/cadastrar_produto.html', {
        'fornecedores': list(fornecedores),
        'categorias': list(categorias),
    })


@login_required
def editar_produto(request, id):
    produto = get_object_or_404(Produto, id=id)
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            nova_quantidade = decimal_ou_none(data.get('quantidade_base')) or Decimal('0')
            with transaction.atomic():
                produto = Produto.objects.select_for_update().get(pk=produto.pk)
                old_preco_custo = produto.preco_custo
                old_preco_venda = produto.preco_venda
                produto.tipo_produto = data['tipo_produto']
                produto.descricao = data['descricao']
                produto.fornecedor_id = data.get('fornecedor_id') or None
                produto.preco_custo = decimal_ou_none(data.get('preco_custo'))
                produto.preco_venda = decimal_ou_none(data.get('preco_venda'))
                produto.estoque_minimo = decimal_ou_none(data.get('estoque_minimo'))
                produto.metros_por_rolo = data.get('metros_por_rolo') or None
                produto.tipo_tinta = data.get('tipo_tinta', 'N/A')
                produto.cor_tinta = data.get('cor_tinta', 'INCOLOR')
                produto.litros_por_vidro = data.get('litros_por_vidro') or None
                produto.unidade_medida = data.get('unidade_medida', 'UN')
                produto.categoria_id = data.get('categoria_id') or None
                preco_custo_mudou = old_preco_custo != produto.preco_custo
                preco_venda_mudou = old_preco_venda != produto.preco_venda
                produto._historico_ja_salvo = True
                produto.save()
                registrar_ajuste_saldo(
                    produto,
                    request.user,
                    nova_quantidade,
                    'Ajuste de saldo na edição do produto',
                )
                produto.refresh_from_db()
                if preco_custo_mudou or preco_venda_mudou:
                    HistoricoPreco.objects.create(
                        produto=produto,
                        preco_custo_antigo=old_preco_custo if preco_custo_mudou else None,
                        preco_custo_novo=produto.preco_custo if preco_custo_mudou else None,
                        preco_venda_antigo=old_preco_venda if preco_venda_mudou else None,
                        preco_venda_novo=produto.preco_venda if preco_venda_mudou else None,
                        usuario=request.user,
                    )
                log_acao(request.user, 'EDITAR', f'Editou produto {produto.descricao}', 'Produto', produto.id)
            return JsonResponse({'ok': True})
        except (json.JSONDecodeError, KeyError, InvalidOperation, TypeError, ValueError):
            return json_erro('Dados inválidos.')
        except ValidationError as exc:
            return json_erro('; '.join(exc.messages))
    fornecedores = Fornecedor.objects.all().values('id', 'nome')
    categorias = Categoria.objects.all().values('id', 'nome')
    return render(request, 'estoque/editar_produto.html', {
        'produto': produto,
        'fornecedores': list(fornecedores),
        'categorias': list(categorias),
        'produto_json': json.dumps({
            'id': produto.id,
            'tipo_produto': produto.tipo_produto,
            'descricao': produto.descricao,
            'fornecedor_id': produto.fornecedor_id or '',
            'categoria_id': produto.categoria_id or '',
            'quantidade_base': float(produto.quantidade_base) if produto.quantidade_base else '',
            'preco_custo': float(produto.preco_custo) if produto.preco_custo is not None else '',
            'preco_venda': float(produto.preco_venda) if produto.preco_venda is not None else '',
            'estoque_minimo': float(produto.estoque_minimo) if produto.estoque_minimo is not None else '',
            'metros_por_rolo': float(produto.metros_por_rolo) if produto.metros_por_rolo else None,
            'tipo_tinta': produto.tipo_tinta,
            'cor_tinta': produto.cor_tinta,
            'litros_por_vidro': float(produto.litros_por_vidro) if produto.litros_por_vidro else None,
            'unidade_medida': produto.unidade_medida or 'UN',
        }),
    })


@login_required
def excluir_produto(request, id):
    produto = get_object_or_404(Produto, id=id)
    if request.method == 'POST':
        perm_error = exigir_admin_json(request)
        if perm_error:
            return perm_error
        if produto.movimentacoes.exists():
            return json_erro('O produto não pode ser excluído porque possui movimentações vinculadas.', codigo='VINCULO_IMPEDITIVO')
        if ItemOrdemCompra.objects.filter(produto=produto).exists():
            return json_erro('O produto não pode ser excluído porque possui ordens de compra vinculadas.', codigo='VINCULO_IMPEDITIVO')
        descricao = produto.descricao
        produto.delete()
        log_acao(request.user, 'EXCLUIR', f'Excluiu produto {descricao}', 'Produto', id)
        return json_ok(mensagem='Produto excluído com sucesso.')
    return json_erro('Exclusão deve usar POST.', status=405)


@login_required
def detalhe_produto(request, id):
    produto = get_object_or_404(Produto, id=id)
    movimentacoes = Movimentacao.objects.filter(produto=produto).select_related('produto').order_by('-data')
    historico_precos = HistoricoPreco.objects.filter(produto=produto)[:20]
    lucro = (produto.preco_venda - produto.preco_custo) if produto.preco_venda is not None and produto.preco_custo is not None else None
    margem = (lucro / produto.preco_custo * 100) if lucro is not None and produto.preco_custo and produto.preco_custo > 0 else None
    return render(request, 'estoque/detalhe.html', {
        'produto': produto,
        'movimentacoes': movimentacoes,
        'historico_precos': historico_precos,
        'lucro': round(lucro, 2) if lucro is not None else None,
        'margem': round(margem, 1) if margem is not None else None,
    })


@login_required
def etiqueta_produto(request, id):
    produto = get_object_or_404(Produto, id=id)
    return render(request, 'estoque/etiqueta.html', {'produto': produto})


@login_required
def template_csv_produtos(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="modelo_importacao_produtos.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'descricao', 'tipo_produto', 'unidade_medida', 'quantidade_base',
        'estoque_minimo', 'preco_custo', 'preco_venda', 'fornecedor_nome',
        'categoria_nome', 'metros_por_rolo', 'tipo_tinta', 'cor_tinta', 'litros_por_vidro'
    ])
    writer.writerow([
        'Exemplo Papel Sublimático', 'PAPEL', 'M', '100',
        '20', '15.50', '25.00', 'Fornecedor A',
        'Papéis', '100', '', '', ''
    ])
    return response


def parse_decimal(val, default=None):
    if not val or not str(val).strip():
        return default
    try:
        return Decimal(str(val).strip().replace(',', '.'))
    except (InvalidOperation, ValueError):
        return default


def parse_decimal_zero(val):
    res = parse_decimal(val, default=Decimal('0'))
    return res if res is not None else Decimal('0')


@login_required
def importar_csv_produtos(request):
    if request.method != 'POST':
        return json_erro('Método não permitido.', status=405)
    
    if 'arquivo' not in request.FILES:
        return json_erro('Nenhum arquivo enviado.')

    file = request.FILES['arquivo']
    if not file.name.endswith('.csv'):
        return json_erro('Formato inválido. O arquivo deve ser um .csv')

    try:
        decoded = file.read().decode('utf-8-sig')
        io_string = StringIO(decoded)
        reader = csv.DictReader(io_string)

        novos = 0
        atualizados = 0

        with transaction.atomic():
            for row in reader:
                desc = row.get('descricao', '').strip()
                if not desc:
                    continue

                forn_nome = row.get('fornecedor_nome', '').strip()
                fornecedor = None
                if forn_nome:
                    fornecedor, _ = Fornecedor.objects.get_or_create(nome=forn_nome)

                cat_nome = row.get('categoria_nome', '').strip()
                categoria = None
                if cat_nome:
                    categoria, _ = Categoria.objects.get_or_create(nome=cat_nome)

                tipo_prod = row.get('tipo_produto', '').strip().upper()
                if tipo_prod not in ['PAPEL', 'TECIDO', 'TINTA', 'AVIAMENTO', 'OUTRO']:
                    tipo_prod = 'OUTRO'

                unid = row.get('unidade_medida', '').strip().upper()
                if not unid:
                    unid = 'UN'

                qtd = parse_decimal_zero(row.get('quantidade_base'))
                minimo = parse_decimal(row.get('estoque_minimo'))
                custo = parse_decimal(row.get('preco_custo'))
                venda = parse_decimal(row.get('preco_venda'))
                metros_rolo = parse_decimal(row.get('metros_por_rolo'))
                litros_vidro = parse_decimal(row.get('litros_por_vidro'))
                
                tipo_tinta = row.get('tipo_tinta', '').strip().upper() or 'N/A'
                cor_tinta = row.get('cor_tinta', '').strip().upper() or 'INCOLOR'

                prod = Produto.objects.select_for_update().filter(descricao=desc).first()
                created = prod is None
                if created:
                    prod = Produto.objects.create(
                        descricao=desc,
                        tipo_produto=tipo_prod,
                        unidade_medida=unid,
                        quantidade_base=Decimal('0'),
                        estoque_minimo=minimo,
                        preco_custo=custo,
                        preco_venda=venda,
                        fornecedor=fornecedor,
                        categoria=categoria,
                        metros_por_rolo=metros_rolo,
                        tipo_tinta=tipo_tinta,
                        cor_tinta=cor_tinta,
                        litros_por_vidro=litros_vidro,
                    )
                else:
                    prod.tipo_produto = tipo_prod
                    prod.unidade_medida = unid
                    prod.estoque_minimo = minimo
                    prod.preco_custo = custo
                    prod.preco_venda = venda
                    prod.fornecedor = fornecedor
                    prod.categoria = categoria
                    prod.metros_por_rolo = metros_rolo
                    prod.tipo_tinta = tipo_tinta
                    prod.cor_tinta = cor_tinta
                    prod.litros_por_vidro = litros_vidro
                    prod.save()

                registrar_ajuste_saldo(
                    prod,
                    request.user,
                    qtd,
                    'Ajuste de saldo via importação CSV de produtos',
                )

                if created:
                    novos += 1
                else:
                    atualizados += 1

        log_acao(request.user, 'IMPORTAR', f'Importou CSV de produtos: {novos} novos, {atualizados} atualizados', 'Produto')
        return json_ok(mensagem=f'Importação concluída! {novos} produto(s) criado(s), {atualizados} atualizado(s).')

    except Exception as e:
        return json_erro(f'Erro ao processar arquivo: {str(e)}')


@login_required
def exportar_atual_xlsx(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    headers = [
        "ID", "Descrição", "Tipo de Produto", "Categoria", "Fornecedor",
        "Qtd Base (Saldo)", "Unidade", "Qtd Formatada", "Estoque Mínimo",
        "Status Estoque", "Preço Custo (R$)", "Preço Venda (R$)",
        "Lucro Unit. (R$)", "Margem (%)", "Valor Total Custo (R$)",
        "Metros/Rolo", "Rolos Est.", "Tipo Tinta", "Cor Tinta", "Litros/Vidro", "Garrafas Est."
    ]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    produtos = Produto.objects.select_related('fornecedor', 'categoria').all().order_by('descricao')

    for row_idx, p in enumerate(produtos, start=2):
        lucro = p.lucro
        margem = p.margem_lucro
        val_total = (p.quantidade_base * p.preco_custo) if p.preco_custo is not None else 0

        row = [
            p.id,
            p.descricao,
            p.get_tipo_produto_display(),
            p.categoria.nome if p.categoria else "-",
            p.fornecedor.nome if p.fornecedor else "-",
            float(p.quantidade_base),
            p.unidade_simbolo,
            p.quantidade_formatada,
            float(p.estoque_minimo) if p.estoque_minimo is not None else "-",
            p.status_estoque,
            float(p.preco_custo) if p.preco_custo is not None else None,
            float(p.preco_venda) if p.preco_venda is not None else None,
            float(lucro) if lucro is not None else None,
            float(margem) if margem is not None else None,
            float(val_total) if p.preco_custo is not None else None,
            float(p.metros_por_rolo) if p.metros_por_rolo else "-",
            p.quantidade_rolos_estimada if p.metros_por_rolo else "-",
            p.get_tipo_tinta_display() if p.tipo_produto == 'TINTA' else "-",
            p.get_cor_tinta_display() if p.tipo_produto == 'TINTA' else "-",
            float(p.litros_por_vidro) if p.litros_por_vidro else "-",
            p.quantidade_vidros_estimada if p.litros_por_vidro else "-",
        ]
        ws.append(row)

        for col_num in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_num)
            c.border = thin_border
            if col_num in [11, 12, 13, 15]:
                c.number_format = 'R$ #,##0.00'
            elif col_num == 14:
                c.number_format = '0.0"%"'
            elif col_num == 6:
                c.number_format = '#,##0.00'

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="produtos_estoque.xlsx"'
    wb.save(response)
    return response


@login_required
def busca_rapida(request):
    q = request.GET.get('q', '').strip()
    if not q or len(q) < 2:
        return JsonResponse({'resultados': []})

    produtos = Produto.objects.filter(
        Q(descricao__icontains=q) |
        Q(fornecedor__nome__icontains=q)
    ).select_related('fornecedor')[:10]

    resultados = []
    for p in produtos:
        resultados.append({
            'id': p.id,
            'descricao': p.descricao,
            'fornecedor': p.fornecedor.nome if p.fornecedor else 'Sem fornecedor',
            'quantidade': p.quantidade_formatada,
            'tipo_produto': p.get_tipo_produto_display(),
        })

    return JsonResponse({'resultados': resultados})
