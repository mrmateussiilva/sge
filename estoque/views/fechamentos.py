import calendar
import json
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..log_utils import log_acao
from ..models import FechamentoMensal
from ..services.fechamentos import criar_fechamento_periodo
from ..services.units import UNIDADES, dinheiro_br
from .helpers import (
    data_iso, exigir_admin_json, json_erro, json_ok, requisicao_htmx, resumo_fechamento
)


@login_required
def lista_fechamentos(request):
    hoje = timezone.localdate()
    data_inicio = hoje.replace(day=1)
    data_fim = hoje.replace(day=calendar.monthrange(hoje.year, hoje.month)[1])
    return render(request, 'estoque/fechamentos.html', {
        **contexto_lista_fechamentos(),
        'data_inicio_sugerida': data_inicio.isoformat(),
        'data_fim_sugerida': data_fim.isoformat(),
        'resumo_inicial': resumo_fechamento(data_inicio, data_fim),
    })


def contexto_lista_fechamentos():
    fechamentos = FechamentoMensal.objects.prefetch_related('itens').select_related('usuario').all()
    fechamentos_data = []
    for f in fechamentos:
        itens = list(f.itens.all())
        total_itens = len(itens)
        valor_total = Decimal('0.00')
        produtos_sem_custo = 0
        for item in itens:
            if item.quantidade > 0 and item.preco_custo is None:
                produtos_sem_custo += 1
            elif item.preco_custo is not None:
                valor_total += item.quantidade * item.preco_custo
        fechamentos_data.append({
            'id': f.id,
            'data_fechamento': f.data_fechamento.strftime('%d/%m/%Y %H:%M'),
            'usuario': f.usuario.username if f.usuario else '-',
            'periodo_formatado': f.periodo_formatado,
            'observacao': f.observacao,
            'total_itens': total_itens,
            'valor_total_formatado': dinheiro_br(valor_total),
            'produtos_sem_custo': produtos_sem_custo,
            'calculo_completo': produtos_sem_custo == 0,
        })
    return {'fechamentos': fechamentos_data}


@login_required
def detalhe_fechamento(request, id):
    fechamento = get_object_or_404(
        FechamentoMensal.objects.select_related('usuario'),
        id=id,
    )
    itens = []
    valor_total = Decimal('0.00')
    valor_total_venda = Decimal('0.00')
    produtos_sem_custo = 0
    for item in fechamento.itens.all().order_by('descricao'):
        valor_custo = None
        if item.preco_custo is None:
            if item.quantidade > 0:
                produtos_sem_custo += 1
        else:
            valor_custo = item.quantidade * item.preco_custo
            valor_total += valor_custo
        valor_venda = None
        if item.preco_venda is not None:
            valor_venda = item.quantidade * item.preco_venda
            valor_total_venda += valor_venda
        unidade = UNIDADES.get(item.unidade_medida, UNIDADES['OUTRO'])
        itens.append({
            'descricao': item.descricao,
            'tipo': item.get_tipo_produto_display() or '—',
            'categoria': item.categoria_nome or '—',
            'fornecedor': item.fornecedor_nome or '—',
            'unidade': unidade.simbolo or '—',
            'quantidade': item.quantidade,
            'preco_custo': item.preco_custo,
            'preco_venda': item.preco_venda,
            'valor_custo': valor_custo,
            'valor_venda': valor_venda,
        })
    return render(request, 'estoque/detalhe_fechamento.html', {
        'fechamento': fechamento,
        'itens': itens,
        'valor_total': valor_total,
        'valor_total_formatado': dinheiro_br(valor_total),
        'valor_total_venda': valor_total_venda,
        'valor_total_venda_formatado': dinheiro_br(valor_total_venda),
        'produtos_sem_custo': produtos_sem_custo,
    })


@login_required
def excluir_fechamento(request, id):
    if request.method != 'POST':
        return json_erro('Exclusão deve usar POST.', status=405)
    perm_error = exigir_admin_json(request)
    if perm_error:
        return perm_error

    with transaction.atomic():
        fechamento = get_object_or_404(
            FechamentoMensal.objects.select_for_update(),
            id=id,
        )
        periodo = fechamento.periodo_formatado
        total_itens = fechamento.itens.count()
        fechamento.delete()
        log_acao(
            request.user,
            'EXCLUIR',
            f'Excluiu fechamento de estoque de {periodo} com {total_itens} item(ns)',
            'FechamentoMensal',
            id,
        )

    if requisicao_htmx(request):
        response = render(
            request,
            'estoque/fechamentos/_lista.html',
            contexto_lista_fechamentos(),
        )
        response['HX-Trigger-After-Swap'] = json.dumps({
            'sge:feedback': {
                'message': 'Fechamento excluído. O período já pode receber um novo freeze.',
                'type': 'success',
            },
        })
        return response
    return json_ok(mensagem='Fechamento excluído. O período já pode receber um novo freeze.')


@login_required
def revisar_fechamento(request):
    try:
        inicio = data_iso(request.GET.get('data_inicio'), 'Data inicial')
        fim = data_iso(request.GET.get('data_fim'), 'Data final')
        resumo = resumo_fechamento(inicio, fim)
        if resumo['duplicado']:
            mensagem = f'Já existe um fechamento para {resumo["periodo_formatado"]}.'
            if requisicao_htmx(request):
                return render(request, 'estoque/fechamentos/_revisao.html', {'erro': mensagem})
            return json_erro(
                mensagem,
                codigo='FECHAMENTO_DUPLICADO',
            )
        if requisicao_htmx(request):
            return render(request, 'estoque/fechamentos/_revisao.html', {'resumo': resumo})
        return json_ok(resumo=resumo)
    except ValidationError as exc:
        if requisicao_htmx(request):
            return render(
                request,
                'estoque/fechamentos/_revisao.html',
                {'erro': '; '.join(exc.messages)},
            )
        return json_erro('; '.join(exc.messages))


@login_required
def realizar_fechamento(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'erro': 'Método não permitido.'}, status=405)
    try:
        data = request.POST if requisicao_htmx(request) else json.loads(request.body)
        inicio = data_iso(data.get('data_inicio'), 'Data inicial')
        fim = data_iso(data.get('data_fim'), 'Data final')
        observacao = (data.get('observacao') or '').strip()
        fechamento = criar_fechamento_periodo(
            data_inicio=inicio,
            data_fim=fim,
            usuario=request.user,
            observacao=observacao,
        )
        if requisicao_htmx(request):
            messages.success(request, 'Fechamento de estoque realizado com sucesso.')
            response = HttpResponse()
            response['HX-Redirect'] = reverse('lista_fechamentos')
            return response
        return json_ok(mensagem='Fechamento de estoque realizado com sucesso.', id=fechamento.id)
    except json.JSONDecodeError:
        return json_erro('JSON inválido.')
    except ValidationError as exc:
        mensagem = '; '.join(exc.messages)
        if requisicao_htmx(request):
            return render(request, 'estoque/fechamentos/_revisao.html', {'erro': mensagem})
        codigo = 'FECHAMENTO_DUPLICADO' if 'Já existe' in mensagem else 'VALIDACAO'
        return json_erro(mensagem, codigo=codigo)


@login_required
def exportar_fechamento_xlsx(request, id):
    fechamento = get_object_or_404(FechamentoMensal, id=id)
    itens = fechamento.itens.all().order_by('descricao')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Fechamento {fechamento.data_fim:%Y-%m-%d}"

    font_title = Font(name='Segoe UI', size=16, bold=True, color='1E293B')
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_data = Font(name='Segoe UI', size=11, color='1E293B')
    font_total = Font(name='Segoe UI', size=11, bold=True, color='1E293B')
    
    fill_header = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    fill_total = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    ws.merge_cells('A1:J1')
    ws['A1'] = f"S.G.E - Fechamento de Estoque ({fechamento.periodo_formatado})"
    ws['A1'].font = font_title
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Realizado em: {fechamento.data_fechamento.strftime('%d/%m/%Y %H:%M')} por {fechamento.usuario.username if fechamento.usuario else '-'}"
    ws['A2'].font = Font(name='Segoe UI', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    headers = [
        "Descrição do Material", "Tipo", "Categoria", "Fornecedor",
        "Unid.", "Quantidade", "Preço Custo",
        "Preço Venda", "Total Custo", "Total Venda"
    ]
    
    ws.append([])
    ws.row_dimensions[4].height = 28
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center' if col_num > 4 else 'left', vertical='center')
        cell.border = border_thin

    start_row = 5
    for item in itens:
        tipo = item.get_tipo_produto_display() or '-'
        fornecedor = item.fornecedor_nome or '-'
        unidade_info_snapshot = UNIDADES.get(item.unidade_medida, UNIDADES['OUTRO'])
        unidade = unidade_info_snapshot.simbolo
        preco_custo = item.preco_custo
        preco_venda = item.preco_venda
        next_row = ws.max_row + 1
        
        row_data = [
            item.descricao,
            tipo,
            item.categoria_nome or '-',
            fornecedor,
            unidade,
            float(item.quantidade),
            float(preco_custo) if preco_custo is not None else None,
            float(preco_venda) if preco_venda is not None else None,
            f"=F{next_row}*G{next_row}" if preco_custo is not None else None,
            f"=F{next_row}*H{next_row}" if preco_venda is not None else None,
        ]
        
        ws.append(row_data)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 20
        
        for col_idx in range(1, 11):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = font_data
            cell.border = border_thin
            
            if col_idx in (5, 6):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in (7, 8, 9, 10):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
            if col_idx in (7, 8, 9, 10):
                cell.number_format = 'R$ #,##0.00'
            elif col_idx == 6:
                cell.number_format = '#,##0.00'

    end_row = ws.max_row
    ws.append([
        "TOTAL GERAL", "", "", "", "",
        "Quantidades por unidade não são somadas", "", "",
        f"=SUM(I{start_row}:I{end_row})",
        f"=SUM(J{start_row}:J{end_row})"
    ])
    
    total_row = ws.max_row
    ws.row_dimensions[total_row].height = 26
    
    for col_idx in range(1, 11):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = font_total
        cell.fill = fill_total
        cell.border = border_thin
        
        if col_idx in (6, 9, 10):
            cell.alignment = Alignment(horizontal='left' if col_idx == 6 else 'right', vertical='center')
            if col_idx in (9, 10):
                cell.number_format = 'R$ #,##0.00'
            else:
                cell.number_format = '#,##0.00'
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.row > 2 and cell.value:
                val_str = str(cell.value)
                if val_str.startswith('='):
                    val_str = "R$ 999.999,99"
                max_len = max(max_len, len(val_str))
                
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename="fechamento_{fechamento.data_inicio:%d-%m-%Y}'
        f'_a_{fechamento.data_fim:%d-%m-%Y}.xlsx"'
    )
    wb.save(response)
    return response
